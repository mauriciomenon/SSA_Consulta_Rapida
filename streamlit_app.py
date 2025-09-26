"""Streamlit frontend otimizado para explorar SSAs utilizando o banco local."""
from __future__ import annotations

import hashlib
import json
import logging
import os
import time
from datetime import datetime
from typing import Dict, Optional, Tuple

import pandas as pd
import streamlit as st

# Inicializar logging robusto
try:
    from utils.robust_logging import setup_logging
    setup_logging()
    logger = logging.getLogger(__name__)
    logger.info("Sistema de logging robusto inicializado no Streamlit", extra={'component': 'streamlit'})
except Exception as e:
    # Fallback para logging padrão
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    logger.error(f"Falha ao inicializar logging robusto: {e}")

from core.app_logic import (
    filter_dataframe,
    get_filtered_data,
    import_files_to_database,
    parse_search_terms,
)
from core.config_manager import load_display_mappings_integrity
from utils.remote_itaipu import RequestOptions, fetch_pending_ssas, map_to_dataframe

DB_PATH_DEFAULT = os.environ.get("SSA_DB_PATH", "data/ssas.db")
DOCS_DIR_DEFAULT = os.environ.get("SSA_DOCS_DIR", "docs_entrada")
DISPLAY_MAPPINGS = load_display_mappings_integrity()


# === Sistema de Cache para Streamlit ===
class StreamlitFilterCache:
    """Cache inteligente para filtros do Streamlit com TTL e estatísticas."""
    
    def __init__(self, max_size: int = 30, ttl_seconds: int = 300):
        self.max_size = max_size
        self.ttl_seconds = ttl_seconds
        
        # Inicializa session state se não existe
        if 'filter_cache' not in st.session_state:
            st.session_state.filter_cache = {}
        if 'cache_stats' not in st.session_state:
            st.session_state.cache_stats = {'hits': 0, 'misses': 0, 'evictions': 0}
    
    def _generate_key(self, df_shape: Tuple[int, int], search_terms: str, 
                     situacoes: list, executores: list, emissores: list) -> str:
        """Gera chave única para o cache baseada nos parâmetros de filtro."""
        params = {
            'shape': df_shape,
            'search': search_terms,
            'situacoes': sorted(situacoes) if situacoes else [],
            'executores': sorted(executores) if executores else [],
            'emissores': sorted(emissores) if emissores else []
        }
        
        params_str = str(sorted(params.items()))
        return hashlib.md5(params_str.encode('utf-8')).hexdigest()
    
    def get(self, df_shape: Tuple[int, int], search_terms: str, 
           situacoes: list, executores: list, emissores: list) -> Optional[pd.DataFrame]:
        """Recupera resultado do cache se válido."""
        key = self._generate_key(df_shape, search_terms, situacoes, executores, emissores)
        
        cache = st.session_state.filter_cache
        if key in cache:
            entry = cache[key]
            # Verifica TTL
            if time.time() - entry['timestamp'] < self.ttl_seconds:
                # Move para o final (LRU)
                cache[key] = cache.pop(key)
                st.session_state.cache_stats['hits'] += 1
                return entry['data'].copy()
            else:
                # Cache expirado
                del cache[key]
        
        st.session_state.cache_stats['misses'] += 1
        return None
    
    def put(self, df_shape: Tuple[int, int], search_terms: str, 
           situacoes: list, executores: list, emissores: list, result: pd.DataFrame):
        """Armazena resultado no cache."""
        key = self._generate_key(df_shape, search_terms, situacoes, executores, emissores)
        
        cache = st.session_state.filter_cache
        
        # Remove entrada existente se houver
        if key in cache:
            del cache[key]
        
        # Implementa política LRU
        while len(cache) >= self.max_size:
            # Remove item mais antigo
            oldest_key = next(iter(cache))
            del cache[oldest_key]
            st.session_state.cache_stats['evictions'] += 1
        
        # Adiciona nova entrada
        cache[key] = {
            'data': result.copy(),
            'timestamp': time.time()
        }
    
    def get_stats(self) -> dict:
        """Retorna estatísticas do cache."""
        stats = st.session_state.cache_stats
        total = stats['hits'] + stats['misses']
        hit_rate = (stats['hits'] / total * 100) if total > 0 else 0
        
        return {
            'size': len(st.session_state.filter_cache),
            'max_size': self.max_size,
            'hits': stats['hits'],
            'misses': stats['misses'],
            'evictions': stats['evictions'],
            'hit_rate': hit_rate,
            'ttl_seconds': self.ttl_seconds
        }
    
    def clear(self):
        """Limpa todo o cache."""
        st.session_state.filter_cache = {}
        st.session_state.cache_stats = {'hits': 0, 'misses': 0, 'evictions': 0}


# Instancia cache global
filter_cache = StreamlitFilterCache()


@st.cache_data(show_spinner=False)
def load_dataframe(db_path: str) -> pd.DataFrame:
    if not os.path.exists(db_path):
        return pd.DataFrame()
    return get_filtered_data(db_path)


def apply_cli_filters(df: pd.DataFrame, search_text: str) -> pd.DataFrame:
    """Aplica filtros CLI com fallback para caso sem cache."""
    if not search_text.strip():
        return df
    raw_terms = [term.strip() for term in search_text.split(',') if term.strip()]
    parsed = parse_search_terms(raw_terms)
    return filter_dataframe(df, parsed)


def apply_all_filters_cached(df: pd.DataFrame, search_terms: str, 
                           situacoes: list, executores: list, emissores: list) -> pd.DataFrame:
    """Aplica todos os filtros com cache inteligente."""
    # Verifica cache primeiro
    cached_result = filter_cache.get(df.shape, search_terms, situacoes, executores, emissores)
    if cached_result is not None:
        return cached_result
    
    # Cache miss - aplica filtros
    start_time = time.time()
    
    # Filtro de busca textual
    filtered_df = apply_cli_filters(df, search_terms)
    
    # Filtros de seleção múltipla
    if situacoes:
        filtered_df = filtered_df[filtered_df['situacao'].isin(situacoes)]
    if executores:
        filtered_df = filtered_df[filtered_df['setor_executor'].isin(executores)]
    if emissores:
        filtered_df = filtered_df[filtered_df['setor_emissor'].isin(emissores)]
    
    # Reset index
    filtered_df = filtered_df.reset_index(drop=True)
    
    # Armazena no cache
    filter_cache.put(df.shape, search_terms, situacoes, executores, emissores, filtered_df)
    
    # Log performance se demorou mais que 100ms
    elapsed = time.time() - start_time
    if elapsed > 0.1:
        st.info(f"⏱️ Filtro executado em {elapsed:.2f}s - resultado armazenado no cache")
    
    return filtered_df


def ensure_arrow_compatible(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza colunas para evitar falhas do Streamlit/Arrow."""
    safe = df.copy()
    for col in safe.columns:
        series = safe[col]
        if series.dtype == "object":
            non_null = series.dropna()
            if not non_null.empty:
                sample_types = {type(x) for x in non_null.head(20)}
                if len(sample_types) > 1:
                    safe[col] = series.astype(str)
                    continue
                sample = non_null.iloc[0]
                if isinstance(sample, (list, dict)):
                    safe[col] = series.apply(
                        lambda x: json.dumps(x, ensure_ascii=False)
                        if isinstance(x, (list, dict))
                        else x
                    )
        elif pd.api.types.is_integer_dtype(series.dtype):
            safe[col] = series.astype("Int64")
    return safe


st.set_page_config(
    page_title="SSA Consulta Rápida", 
    layout="wide",
    initial_sidebar_state="expanded",
    page_icon="⚡"
)

# Header com indicadores de status
col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.title("⚡ SSA Consulta Rápida – Dashboard Otimizado")
with col2:
    # Indicador de cache
    cache_stats = filter_cache.get_stats()
    if cache_stats['hits'] + cache_stats['misses'] > 0:
        cache_color = "🟢" if cache_stats['hit_rate'] > 80 else "🟡" if cache_stats['hit_rate'] > 50 else "🔴"
        st.metric("Cache", f"{cache_color} {cache_stats['hit_rate']:.0f}%", help="Taxa de acertos do cache")
with col3:
    # Timestamp da última atualização
    st.metric("Atualizado", datetime.now().strftime("%H:%M:%S"), help="Última atualização da página")

with st.sidebar:
    with st.expander("Fonte de dados", expanded=True):
        db_path = st.text_input("Caminho do banco SQLite", value=DB_PATH_DEFAULT)
        docs_dir = st.text_input("Diretório de planilhas", value=DOCS_DIR_DEFAULT)
        col_sync_1, col_sync_2 = st.columns(2)
        with col_sync_1:
            force_import = st.checkbox("Forçar reimportação", value=False)
        with col_sync_2:
            if st.button("Atualizar banco", use_container_width=True):
                import_start = time.time()
                
                # Progress bar personalizada
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    status_text.text("🔍 Verificando arquivos...")
                    progress_bar.progress(20)
                    
                    status_text.text("📊 Importando planilhas...")
                    progress_bar.progress(50)
                    
                    ok = import_files_to_database(
                        docs_dir=docs_dir,
                        db_path=db_path,
                        force_import=force_import,
                    )
                    
                    progress_bar.progress(80)
                    status_text.text("🧹 Limpando cache...")
                    
                    # Limpa caches após importação
                    load_dataframe.clear()
                    filter_cache.clear()
                    
                    progress_bar.progress(100)
                    
                    elapsed = time.time() - import_start
                    
                    if ok:
                        status_text.text(f"✅ Importação concluída em {elapsed:.1f}s")
                        st.success(f"Dados importados com sucesso! ({elapsed:.1f}s)")
                    else:
                        status_text.text("❌ Falha na importação")
                        st.error("Falha ao importar dados. Verifique os logs para detalhes.")
                
                except Exception as e:
                    progress_bar.progress(0)
                    status_text.text("❌ Erro durante importação")
                    st.error(f"Erro durante importação: {e}")
                
                finally:
                    # Remove progress bar após 2 segundos
                    time.sleep(2)
                    progress_bar.empty()
                    status_text.empty()

    with st.expander("Filtros", expanded=True):
        search_terms = st.text_input(
            "Busca rapida (virgulas, mesma sintaxe da CLI)",
            value="",
            help="Separe por virgulas. Use OU/OR para alternativas e ! para exclusoes.",
        )
        col1, col2 = st.columns(2)
        with col1:
            limit_rows = st.slider("Limitar linhas", min_value=50, max_value=5000, value=500, step=50)
        with col2:
            show_progress = st.checkbox("Progress bars", value=True, help="Mostra indicadores de progresso durante operações")
        
        consult_api = st.checkbox(
            "Consultar API Itaipu por dados recentes",
            value=False,
            help=(
                "A API oficial fornece apenas campos básicos (número, situação, setores, emissor). "
                "Se estiver indisponível, os dados locais continuam sendo exibidos normalmente."
            ),
        )

    with st.expander("⚡ Performance & Cache", expanded=False):
        cache_stats = filter_cache.get_stats()
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Cache Hits", cache_stats['hits'])
            st.metric("Hit Rate", f"{cache_stats['hit_rate']:.1f}%")
        with col2:
            st.metric("Cache Size", f"{cache_stats['size']}/{cache_stats['max_size']}")
            st.metric("TTL", f"{cache_stats['ttl_seconds']}s")
        
        if st.button("🗑️ Limpar Cache", help="Remove todos os filtros do cache"):
            filter_cache.clear()
            st.success("Cache limpo!")
            st.rerun()
        
        if cache_stats['hits'] + cache_stats['misses'] > 0:
            if cache_stats['hit_rate'] > 80:
                st.success("🚀 Cache performance: EXCELENTE")
            elif cache_stats['hit_rate'] > 50:
                st.info("⚡ Cache performance: BOM")
            else:
                st.warning("⏳ Cache performance: REGULAR")

# Carregar dados
raw_df = load_dataframe(db_path)
if raw_df.empty:
    st.info(
        "Banco não encontrado ou sem dados. Utilize o botão 'Atualizar banco' na barra lateral "
        "para importar as planilhas."
    )
    st.stop()

# Filtros adicionais (exibidos com base no dataset carregado)
with st.sidebar:
    st.subheader("Filtros rápidos")
    situacoes = sorted([s for s in raw_df.get('situacao', pd.Series(dtype=str)).dropna().unique()])
    situacao_sel = st.multiselect("Situações", situacoes, default=situacoes[:1])
    executores = sorted([s for s in raw_df.get('setor_executor', pd.Series(dtype=str)).dropna().unique()])
    default_executor = ['IEE3'] if 'IEE3' in executores else executores[:1]
    executor_sel = st.multiselect("Setores executores", executores, default=default_executor)
    emissores = sorted([s for s in raw_df.get('setor_emissor', pd.Series(dtype=str)).dropna().unique()])
    default_emissor = ['IEE3'] if 'IEE3' in emissores else emissores[:1]
    emissor_sel = st.multiselect("Setores emissores", emissores, default=default_emissor)
    column_display_names = {
        col: DISPLAY_MAPPINGS.get(col, col)
        for col in raw_df.columns
    }
    default_columns = [col for col in raw_df.columns if col in (
        'numero_ssa', 'situacao', 'descricao_ssa', 'setor_executor', 'setor_emissor',
        'data_cadastro', 'prazo_limite'
    )]
    if not default_columns:
        default_columns = list(raw_df.columns[:10])
    selected_display = st.multiselect(
        "Colunas para exibir",
        options=[column_display_names[col] for col in raw_df.columns],
        default=[column_display_names[col] for col in default_columns],
    )
    display_to_internal = {v: k for k, v in column_display_names.items()}
    selected_columns = [display_to_internal.get(name, name) for name in selected_display]

    with st.expander("Ajuda de filtros", expanded=False):
        st.markdown(
            """
* Sintaxe basica: `svp, !ste, mel4`
* Use `OU` ou `OR` para alternativas (`svp OU mel4`)
* Prefixos uteis: `^` inicio, `$` final, `=` igual, `~` regex
* `!` inverte um termo (`!^adm`, `!mel4`)
* Virgulas equivalem a E/AND; espacos tambem separam termos
            """
        )


# Aplica todos os filtros com cache inteligente
filtered_df = apply_all_filters_cached(
    raw_df, 
    search_terms, 
    situacao_sel, 
    executor_sel, 
    emissor_sel
)
if limit_rows and len(filtered_df) > limit_rows:
    filtered_df = filtered_df.head(limit_rows)

active_summary: list[str] = []
if search_terms.strip():
    active_summary.append(f"Busca: {search_terms.strip()}")
if situacao_sel and situacoes and len(situacao_sel) != len(situacoes):
    active_summary.append("Situacao: " + ", ".join(situacao_sel))
if executor_sel and executores and len(executor_sel) != len(executores):
    active_summary.append("Executor: " + ", ".join(executor_sel))
if emissor_sel and emissores and len(emissor_sel) != len(emissores):
    active_summary.append("Emissor: " + ", ".join(emissor_sel))
if consult_api:
    active_summary.append("API: ativada")

if active_summary:
    st.markdown("**Filtros ativos:** " + " | ".join(active_summary))

# Indicadores rápidos aprimorados
metric_cols = st.columns(4)

# Total de SSAs com indicador de performance
total_ssas = len(filtered_df)
original_count = len(raw_df)
reduction_pct = ((original_count - total_ssas) / original_count * 100) if original_count > 0 else 0

metric_cols[0].metric(
    "Total Filtrado", 
    total_ssas,
    delta=f"-{reduction_pct:.1f}%" if reduction_pct > 0 else None,
    help=f"De {original_count} SSAs originais"
)

if 'situacao' in filtered_df.columns:
    status_counts = filtered_df['situacao'].value_counts()
    
    executadas = int(status_counts.get('EXECUTADA', 0))
    metric_cols[1].metric("Executadas", executadas)
    
    pendentes = int(status_counts.get('ABERTA', 0))
    metric_cols[2].metric("Pendentes", pendentes)
    
    # Taxa de execução
    exec_rate = (executadas / total_ssas * 100) if total_ssas > 0 else 0
    metric_cols[3].metric("Taxa Execução", f"{exec_rate:.1f}%")
else:
    metric_cols[1].metric("Executadas", "-")
    metric_cols[2].metric("Pendentes", "-")
    metric_cols[3].metric("Taxa Execução", "-")

# Consulta opcional da API para dados mais recentes (não bloqueia fluxo offline)
recent_df: pd.DataFrame | None = None
if consult_api:
    try:
        api_items = fetch_pending_ssas(years=1, opts=RequestOptions(timeout=5.0))
        mapped = map_to_dataframe(api_items)
        if mapped is not None and not mapped.empty:
            cols = [
                col
                for col in (
                    "numero_ssa",
                    "situacao",
                    "setor_emissor",
                    "setor_executor",
                    "descricao_ssa",
                    "data_cadastro",
                )
                if col in mapped.columns
            ]
            recent_df = mapped[cols].copy()
            st.success(
                f"API Itaipu retornou {len(recent_df)} registros recentes (campos limitados)."
            )
        else:
            st.info("API Itaipu respondeu sem novos registros. Exibindo apenas dados do banco local.")
    except Exception:
            st.warning(
                "Não foi possível acessar dados mais recentes via API. O dashboard continua com o banco local."
            )

# Seção de dados com controles avançados
col1, col2, col3 = st.columns([2, 1, 1])
with col1:
    st.subheader(f"📊 Dados Filtrados ({len(filtered_df)} registros)")
with col2:
    table_height = st.selectbox("Altura da tabela", [400, 600, 800, 1000], index=1, help="Altura em pixels")
with col3:
    auto_width = st.checkbox("Auto width", value=True, help="Ajuste automático de largura das colunas")

# Preparação dos dados para exibição
view_df = filtered_df[selected_columns] if selected_columns else filtered_df
rename_map = {col: DISPLAY_MAPPINGS.get(col, col) for col in view_df.columns}
display_df = ensure_arrow_compatible(view_df.rename(columns=rename_map))

# Configuração avançada de colunas
column_config = {}
for col in view_df.columns:
    display_name = rename_map.get(col, col)
    
    if col in {"situacao", "setor_executor", "setor_emissor"}:
        column_config[display_name] = st.column_config.TextColumn(width="small")
    elif col == "numero_ssa":
        column_config[display_name] = st.column_config.TextColumn(width="medium", help="Número da SSA")
    elif "data" in col.lower():
        column_config[display_name] = st.column_config.DatetimeColumn(width="small")
    elif col == "descricao_ssa":
        column_config[display_name] = st.column_config.TextColumn(width="large")

# Exibe a tabela com configurações otimizadas
if auto_width:
    st.dataframe(
        display_df,
        width='stretch',
        height=table_height,
        column_config=column_config,
        use_container_width=True,
        hide_index=True
    )
else:
    st.dataframe(
        display_df,
        height=table_height,
        column_config=column_config,
        use_container_width=False,
        hide_index=True
    )

# Seção de exportação melhorada
st.markdown("### 📤 Exportação")
col1, col2, col3, col4 = st.columns(4)

with col1:
    # CSV básico
    csv_data = view_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "📄 Baixar CSV",
        csv_data,
        file_name=f"ssas_filtradas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        mime="text/csv",
        help=f"Exporta {len(view_df)} registros em CSV"
    )

with col2:
    # Excel com formatação
    if st.button("📊 Gerar Excel", help="Prepara arquivo Excel com formatação"):
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "SSAs Filtradas"
            
            # Cabeçalhos com formatação
            for col_num, col_name in enumerate(view_df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=rename_map.get(col_name, col_name))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Dados
            for row_num, row_data in enumerate(view_df.itertuples(index=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                "📊 Excel Formatado",
                buffer.getvalue(),
                file_name=f"ssas_filtradas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except ImportError:
            st.error("openpyxl não instalado - usando CSV simples")

with col3:
    # JSON para APIs
    json_data = view_df.to_json(orient='records', date_format='iso', indent=2).encode('utf-8')
    st.download_button(
        "🔗 JSON API",
        json_data,
        file_name=f"ssas_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json",
        help="Formato JSON para integração com APIs"
    )

with col4:
    # Estatísticas do filtro
    if st.button("📈 Estatísticas", help="Mostra resumo estatístico"):
        stats_info = {
            "total_registros": len(view_df),
            "colunas_selecionadas": len(selected_columns),
            "filtros_ativos": len([x for x in [search_terms, situacao_sel, executor_sel, emissor_sel] if x]),
            "cache_hit_rate": f"{filter_cache.get_stats()['hit_rate']:.1f}%"
        }
        st.json(stats_info)

# Gráfico simples de situações
if 'situacao' in filtered_df.columns and not filtered_df.empty:
    chart_df = (
        filtered_df['situacao']
        .value_counts()
        .rename_axis('Situação')
        .reset_index(name='Quantidade')
    )
    st.subheader("Distribuição por Situação")
    st.bar_chart(chart_df.set_index('Situação'))

if recent_df is not None and not recent_df.empty:
    st.markdown("### Dados recentes via API (campos limitados)")
    st.dataframe(ensure_arrow_compatible(recent_df), width='stretch', height=220)
