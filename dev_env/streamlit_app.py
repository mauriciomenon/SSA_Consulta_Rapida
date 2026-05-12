"""Streamlit frontend otimizado para explorar SSAs utilizando o banco local."""

# Last modified: 2025-10-29T11:45:00 (improved arrow compatibility)
from __future__ import annotations

import hashlib
import importlib
import json
import logging
import math
import os
import sqlite3
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional, Tuple, cast

import pandas as pd


def _get_project_root() -> Path:
    return Path(__file__).resolve().parents[1]


project_root = _get_project_root()
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

app_logic = importlib.import_module("core.app_logic")
config_manager = importlib.import_module("core.config_manager")
width_manager_module = importlib.import_module("gui.simple_width_manager")
path_safety_module = importlib.import_module("utils.path_safety")
remote_itaipu_module = importlib.import_module("utils.remote_itaipu")
date_utils_module = importlib.import_module("shared.date_utils")
version_module = importlib.import_module("utils.version")

filter_dataframe = app_logic.filter_dataframe
get_filtered_data = app_logic.get_filtered_data
import_files_to_database = app_logic.import_files_to_database
parse_search_terms = app_logic.parse_search_terms
load_display_mappings_integrity = config_manager.load_display_mappings_integrity
SimpleWidthManager = width_manager_module.SimpleWidthManager
PathSafetyError = path_safety_module.PathSafetyError
ensure_path_is_allowed = path_safety_module.ensure_path_is_allowed
RequestOptions = remote_itaipu_module.RequestOptions
fetch_pending_ssas = remote_itaipu_module.fetch_pending_ssas
map_to_dataframe = remote_itaipu_module.map_to_dataframe
parse_any_date = date_utils_module.parse_any_date
get_app_version = version_module.get_app_version

try:
    st = cast(Any, importlib.import_module("streamlit"))
except ModuleNotFoundError:

    class _StreamlitStub:
        session_state = None

        def __getattr__(self, _name: str):
            return None

    st = cast(Any, _StreamlitStub())

# pandas >= 3 keeps copy-on-write always enabled.


# Inicializar logging robusto
class _ASCIIOnlyFilter(logging.Filter):
    @staticmethod
    def _to_ascii(value):
        if isinstance(value, str):
            return value.encode("ascii", "ignore").decode("ascii")
        return value

    def filter(self, record: logging.LogRecord) -> bool:
        record.msg = self._to_ascii(record.msg)
        if record.args:
            if isinstance(record.args, dict):
                record.args = {
                    key: self._to_ascii(value) for key, value in record.args.items()
                }
            else:
                record.args = tuple(self._to_ascii(arg) for arg in record.args)
        if record.exc_text:
            record.exc_text = self._to_ascii(record.exc_text)
        return True


try:
    from utils.robust_logging import setup_logging

    setup_logging()
except Exception:
    logging.basicConfig(level=logging.INFO)

logger = logging.getLogger(__name__)
ascii_filter = _ASCIIOnlyFilter()
for handler in logging.getLogger().handlers:
    handler.addFilter(ascii_filter)
logger.debug("Logging configurado para Streamlit", extra={"component": "streamlit"})

DB_PATH_DEFAULT = os.environ.get("SSA_DB_PATH", "data/ssas.db")
DOCS_DIR_DEFAULT = os.environ.get("SSA_DOCS_DIR", "docs_entrada")
DISPLAY_MAPPINGS = load_display_mappings_integrity()
APP_VERSION = get_app_version(str(project_root))
STREAMLIT_APP_TITLE = f"SSA Consulta Rapida v{APP_VERSION}"


# === Compatibilidade e Cache para Streamlit ===
# Variaveis padrao para evitar referencias nao definidas quando nao estiver em runtime real
db_path: str = DB_PATH_DEFAULT
docs_dir: str = DOCS_DIR_DEFAULT
search_terms: str = ""
limit_rows: int = 500
consult_api: bool = False
situacoes: list[str] = []
situacao_sel: list[str] = []
executores: list[str] = []
executor_sel: list[str] = []
emissores: list[str] = []
emissor_sel: list[str] = []
selected_columns: list[str] = []
# Placeholders para dados de exibicao quando fora de runtime real
view_df: pd.DataFrame = pd.DataFrame()
rename_map: dict[str, str] = {}
display_df: pd.DataFrame = pd.DataFrame()
column_config: dict[str, object] = {}


def _resolve_api_timeout_seconds() -> float:
    raw = os.environ.get("SSA_STREAMLIT_API_TIMEOUT_SECONDS", "").strip()
    if not raw:
        return 12.0
    try:
        timeout_value = float(raw)
    except ValueError:
        logger.warning("Invalid SSA_STREAMLIT_API_TIMEOUT_SECONDS value: %r", raw)
        return 12.0
    if not math.isfinite(timeout_value) or timeout_value <= 0:
        logger.warning(
            "Invalid SSA_STREAMLIT_API_TIMEOUT_SECONDS non-positive/non-finite: %r", raw
        )
        return 12.0
    return max(2.0, timeout_value)


def _resolve_cache_max_entry_bytes() -> Optional[int]:
    raw = os.environ.get("SSA_CACHE_MAX_MB", "").strip()
    if not raw:
        return None
    try:
        max_mb = float(raw)
    except ValueError:
        logger.warning("Invalid SSA_CACHE_MAX_MB value: %r", raw)
        return None
    if max_mb <= 0:
        return None
    return int(max_mb * 1024 * 1024)


class StreamlitFilterCache:
    """Cache inteligente para filtros do Streamlit com TTL e estatisticas."""

    def __init__(self, max_size: int = 30, ttl_seconds: int = 300):
        self.max_size = max_size
        self.ttl_seconds = ttl_seconds
        self._max_entry_bytes = _resolve_cache_max_entry_bytes()
        self._use_session_state = False
        self._local_cache: dict[str, dict[str, Any]] = {}
        self._local_stats: dict[str, int] = {
            "hits": 0,
            "misses": 0,
            "evictions": 0,
            "skipped_large_entries": 0,
        }
        self._initialize_backend()

    def _initialize_backend(self) -> None:
        if not hasattr(st, "session_state"):
            return
        try:
            session_state = st.session_state
            if session_state is None:
                return
            if "filter_cache" not in session_state:
                session_state.filter_cache = {}
            if "cache_stats" not in session_state:
                session_state.cache_stats = {
                    "hits": 0,
                    "misses": 0,
                    "evictions": 0,
                    "skipped_large_entries": 0,
                }
            session_state.cache_stats.setdefault("skipped_large_entries", 0)
            self._use_session_state = True
        except Exception as exc:
            logger.debug(
                "Streamlit session_state unavailable in current runtime: %s", exc
            )

    def _resolve_backend(self) -> tuple[dict[str, Any], dict[str, int]]:
        if self._use_session_state:
            stats = st.session_state.cache_stats
            stats.setdefault("skipped_large_entries", 0)
            return st.session_state.filter_cache, stats
        self._local_stats.setdefault("skipped_large_entries", 0)
        return self._local_cache, self._local_stats

    def _entry_too_large(self, result: pd.DataFrame) -> bool:
        if self._max_entry_bytes is None:
            return False
        entry_bytes = int(result.memory_usage(index=True, deep=True).sum())
        return entry_bytes > self._max_entry_bytes

    def _get_by_key(self, key: str) -> Optional[pd.DataFrame]:
        cache, stats = self._resolve_backend()
        entry = cache.get(key)
        if not entry:
            stats["misses"] += 1
            return None
        if time.time() - entry["timestamp"] >= self.ttl_seconds:
            del cache[key]
            stats["misses"] += 1
            return None
        cache[key] = cache.pop(key)
        stats["hits"] += 1
        return entry["data"].copy()

    def _store_by_key(
        self, key: str, result: pd.DataFrame, meta: Optional[dict[str, Any]] = None
    ) -> bool:
        cache, stats = self._resolve_backend()
        if self._entry_too_large(result):
            stats["skipped_large_entries"] += 1
            return False
        if key in cache:
            del cache[key]
        while len(cache) >= self.max_size:
            oldest_key = next(iter(cache))
            del cache[oldest_key]
            stats["evictions"] += 1
        cache[key] = {
            "data": result.copy(),
            "timestamp": time.time(),
            "meta": meta or {},
        }
        return True

    def _generate_key(
        self,
        df_shape: Tuple[int, int],
        search_terms: str,
        situacoes: list,
        executores: list,
        emissores: list,
        advanced_filters: Optional[dict[str, Any]] = None,
        df_token: Any = None,
    ) -> str:
        """Gera chave unica para o cache baseada nos parametros de filtro."""
        params = {
            "shape": df_shape,
            "search": search_terms,
            "situacoes": sorted(situacoes) if situacoes else [],
            "executores": sorted(executores) if executores else [],
            "emissores": sorted(emissores) if emissores else [],
            "advanced_filters": _normalize_cache_value(advanced_filters or {}),
            "df_token": df_token,
        }

        params_str = str(sorted(params.items()))
        return hashlib.blake2b(params_str.encode("utf-8"), digest_size=16).hexdigest()

    def get(
        self,
        df_shape: Tuple[int, int],
        search_terms: str,
        situacoes: list,
        executores: list,
        emissores: list,
        advanced_filters: Optional[dict[str, Any]] = None,
        df_token: Any = None,
    ) -> Optional[pd.DataFrame]:
        """Recupera resultado do cache se valido."""
        key = self._generate_key(
            df_shape,
            search_terms,
            situacoes,
            executores,
            emissores,
            advanced_filters=advanced_filters,
            df_token=df_token,
        )
        return self._get_by_key(key)

    def put(
        self,
        df_shape: Tuple[int, int],
        search_terms: str,
        situacoes: list,
        executores: list,
        emissores: list,
        result: pd.DataFrame,
        advanced_filters: Optional[dict[str, Any]] = None,
        df_token: Any = None,
    ):
        """Armazena resultado no cache."""
        key = self._generate_key(
            df_shape,
            search_terms,
            situacoes,
            executores,
            emissores,
            advanced_filters=advanced_filters,
            df_token=df_token,
        )
        stored = self._store_by_key(key, result=result)
        if not stored:
            logger.info(
                "StreamlitFilterCache.put skipped large DataFrame entry (max_bytes=%s)",
                self._max_entry_bytes,
            )

    def get_stats(self) -> dict:
        """Retorna estatisticas do cache."""
        cache, stats = self._resolve_backend()
        total = stats["hits"] + stats["misses"]
        hit_rate = (stats["hits"] / total * 100) if total > 0 else 0

        return {
            "size": len(cache),
            "entries": len(cache),
            "max_size": self.max_size,
            "hits": stats["hits"],
            "misses": stats["misses"],
            "evictions": stats["evictions"],
            "skipped_large_entries": stats["skipped_large_entries"],
            "max_entry_mb": (
                round(self._max_entry_bytes / (1024 * 1024), 3)
                if self._max_entry_bytes is not None
                else None
            ),
            "hit_rate": hit_rate,
            "ttl_seconds": self.ttl_seconds,
        }

    def clear(self):
        """Limpa todo o cache."""
        if self._use_session_state:
            st.session_state.filter_cache = {}
            st.session_state.cache_stats = {
                "hits": 0,
                "misses": 0,
                "evictions": 0,
                "skipped_large_entries": 0,
            }
        else:
            self._local_cache = {}
            self._local_stats = {
                "hits": 0,
                "misses": 0,
                "evictions": 0,
                "skipped_large_entries": 0,
            }

    # --- Metodos de compatibilidade com scripts de teste ---
    def get_cached_filter(self, key: str) -> Optional[pd.DataFrame]:
        return self._get_by_key(key)

    def cache_filter_result(
        self, key: str, result: pd.DataFrame, meta: Optional[dict] = None
    ):
        stored = self._store_by_key(key, result=result, meta=meta)
        if not stored:
            logger.info(
                "StreamlitFilterCache.cache_filter_result skipped large DataFrame entry (max_bytes=%s)",
                self._max_entry_bytes,
            )


# Instancia cache global
filter_cache = StreamlitFilterCache()
width_manager = SimpleWidthManager()
MAX_RENDER_TELEMETRY_PROFILES = 12
WIDTH_PROFILE_OPTIONS = [
    "Compacto (1200)",
    "Padrao (1600)",
    "Largo (2000)",
    "XL (2400)",
]
WIDTH_PROFILE_PIXELS = {
    "Compacto (1200)": 1200,
    "Padrao (1600)": 1600,
    "Largo (2000)": 2000,
    "XL (2400)": 2400,
}
MAIN_SECTION_OPTIONS: list[tuple[str, str]] = [
    ("filtros", "Filtros"),
    ("tabela", "Tabela"),
    ("exportacao", "Exportacao"),
    ("cache_api", "Cache e API"),
]
MAIN_TAB_LABELS = [label for _key, label in MAIN_SECTION_OPTIONS]
STREAMLIT_UI_STATE_FILE_DEFAULT = "streamlit_ui_state.json"
DEFAULT_STREAMLIT_THEME = "GitHub Claro"
STREAMLIT_THEME_PALETTES: dict[str, dict[str, str]] = {
    "GitHub Claro": {
        "bg": "#f6f8fa",
        "panel": "#ffffff",
        "ink": "#24292f",
        "muted": "#57606a",
        "accent": "#0969da",
        "accent_soft": "#ddf4ff",
        "border": "#d0d7de",
        "metric_bg": "#f6f8fa",
        "input_bg": "#ffffff",
        "input_ink": "#24292f",
    },
    "GitHub Escuro": {
        "bg": "#0d1117",
        "panel": "#161b22",
        "ink": "#e6edf3",
        "muted": "#8b949e",
        "accent": "#2f81f7",
        "accent_soft": "#1f2937",
        "border": "#30363d",
        "metric_bg": "#161b22",
        "input_bg": "#0d1117",
        "input_ink": "#e6edf3",
    },
    "Monokai": {
        "bg": "#272822",
        "panel": "#2d2e27",
        "ink": "#f8f8f2",
        "muted": "#b3b3a7",
        "accent": "#a6e22e",
        "accent_soft": "#3e3d32",
        "border": "#5a5a4f",
        "metric_bg": "#2d2e27",
        "input_bg": "#1e1f1c",
        "input_ink": "#f8f8f2",
    },
    "Dracula": {
        "bg": "#282a36",
        "panel": "#343746",
        "ink": "#f8f8f2",
        "muted": "#c0c3ce",
        "accent": "#bd93f9",
        "accent_soft": "#44475a",
        "border": "#6272a4",
        "metric_bg": "#343746",
        "input_bg": "#1f2130",
        "input_ink": "#f8f8f2",
    },
    "Gruvbox Claro": {
        "bg": "#fbf1c7",
        "panel": "#f9f5d7",
        "ink": "#3c3836",
        "muted": "#665c54",
        "accent": "#b57614",
        "accent_soft": "#f2e5bc",
        "border": "#d5c4a1",
        "metric_bg": "#f2e5bc",
        "input_bg": "#fff8dc",
        "input_ink": "#3c3836",
    },
    "Gruvbox Escuro": {
        "bg": "#282828",
        "panel": "#32302f",
        "ink": "#ebdbb2",
        "muted": "#a89984",
        "accent": "#fabd2f",
        "accent_soft": "#3c3836",
        "border": "#504945",
        "metric_bg": "#32302f",
        "input_bg": "#1d2021",
        "input_ink": "#ebdbb2",
    },
    "Windows": {
        "bg": "#f3f3f3",
        "panel": "#ffffff",
        "ink": "#1f1f1f",
        "muted": "#525252",
        "accent": "#0078d4",
        "accent_soft": "#deecf9",
        "border": "#c7c7c7",
        "metric_bg": "#f3f3f3",
        "input_bg": "#ffffff",
        "input_ink": "#1f1f1f",
    },
    "Adwaita": {
        "bg": "#f6f5f4",
        "panel": "#ffffff",
        "ink": "#2e3436",
        "muted": "#5c616c",
        "accent": "#3584e4",
        "accent_soft": "#dce9fb",
        "border": "#c0bfbc",
        "metric_bg": "#f0efee",
        "input_bg": "#ffffff",
        "input_ink": "#2e3436",
    },
    "Atlantico": {
        "bg": "#f2f6fb",
        "panel": "#ffffff",
        "ink": "#10243f",
        "muted": "#52657e",
        "accent": "#1f7a8c",
        "accent_soft": "#dff3f7",
        "border": "#d3dfec",
        "metric_bg": "#f8fbff",
    },
    "Grafite": {
        "bg": "#eef0f3",
        "panel": "#fcfcfd",
        "ink": "#1b1d21",
        "muted": "#4a5059",
        "accent": "#4a5d73",
        "accent_soft": "#e3e9f0",
        "border": "#d1d7df",
        "metric_bg": "#f4f6f8",
    },
    "Solar": {
        "bg": "#fff8ef",
        "panel": "#fffdf8",
        "ink": "#3b2a1f",
        "muted": "#6d5542",
        "accent": "#b95d2a",
        "accent_soft": "#ffe7d6",
        "border": "#ecd6c4",
        "metric_bg": "#fff3e7",
    },
}


def load_dataframe(db_path: str) -> pd.DataFrame:
    """
    Carrega dados do banco de dados com tratamento de erros.

    Returns:
        DataFrame com dados ou DataFrame vazio em caso de erro
    """
    if not os.path.exists(db_path):
        logger.warning(f"Database file not found: {db_path}")
        return pd.DataFrame()

    try:
        df = get_filtered_data(db_path)
        if df.empty:
            logger.info(f"Database query returned empty result: {db_path}")
        else:
            logger.debug(f"Loaded {len(df)} records from {db_path}")
        return df
    except sqlite3.Error as e:
        logger.error(f"SQLite error loading data from {db_path}: {e}")
        return pd.DataFrame()
    except Exception as e:
        logger.error(f"Unexpected error loading data from {db_path}: {e}")
        return pd.DataFrame()


# Aplica cache do Streamlit se disponivel; adiciona clear() no fallback
if hasattr(st, "cache_data") and callable(getattr(st, "cache_data")):
    load_dataframe = st.cache_data(show_spinner=False)(load_dataframe)
else:
    setattr(load_dataframe, "clear", lambda: None)


# Alias para compatibilidade: funcao utilitaria esperada pelos testes
def apply_filters_with_cache(
    df: pd.DataFrame,
    search_terms: str,
    situacoes: list,
    executores: list,
    emissores: list,
    advanced_filters: Optional[dict[str, Any]] = None,
    base_derivada_context_map: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    return apply_all_filters_cached(
        df,
        search_terms,
        situacoes,
        executores,
        emissores,
        advanced_filters=advanced_filters,
        base_derivada_context_map=base_derivada_context_map,
    )


def _compute_df_cache_token(df: pd.DataFrame) -> tuple[Any, ...]:
    """Compute lightweight token to reduce stale cache hits on equal shapes."""
    cached_token = df.attrs.get("_streamlit_cache_token")
    if cached_token is not None:
        return cached_token

    columns = tuple(str(col) for col in df.columns)
    if len(df.columns) == 0:
        token = (len(df), columns, None, None)
        df.attrs["_streamlit_cache_token"] = token
        return token
    if df.empty:
        token = (0, columns, None, None)
        df.attrs["_streamlit_cache_token"] = token
        return token
    sample_column = "numero_ssa" if "numero_ssa" in df.columns else df.columns[0]
    sample_series = df[sample_column]
    head_values = tuple(str(value) for value in sample_series.head(10).tolist())
    tail_values = tuple(str(value) for value in sample_series.tail(10).tolist())
    token = (len(df), columns, str(sample_column), head_values, tail_values)
    df.attrs["_streamlit_cache_token"] = token
    return token


def _build_streamlit_column_config(
    page_df: pd.DataFrame,
    rename_map: dict[str, str],
    available_width: int = 1600,
) -> dict[str, Any]:
    column_config_api = getattr(st, "column_config", None)
    if column_config_api is None:
        return {}
    computed_buckets = width_manager.compute_streamlit_width_buckets(
        page_df,
        available_width=available_width,
        column_order=list(page_df.columns),
    )
    config: dict[str, Any] = {}
    for col in page_df.columns:
        display_name = rename_map.get(col, col)
        bucket = str(computed_buckets.get(col, "medium"))
        if "data" in col.lower():
            config[display_name] = column_config_api.DatetimeColumn(width=bucket)
        elif col == "numero_ssa":
            config[display_name] = column_config_api.TextColumn(
                width=bucket,
                help="Numero da SSA",
            )
        else:
            config[display_name] = column_config_api.TextColumn(width=bucket)
    return config


def _update_render_telemetry(width_profile: str, render_ms: float) -> None:
    if not hasattr(st, "session_state") or st.session_state is None:
        return
    render_stats = st.session_state.get("streamlit_render_stats", {})
    profile_stats = render_stats.get(
        width_profile, {"count": 0, "total_ms": 0.0, "last_ms": 0.0}
    )
    profile_stats["count"] = int(profile_stats.get("count", 0)) + 1
    profile_stats["total_ms"] = float(profile_stats.get("total_ms", 0.0)) + render_ms
    profile_stats["last_ms"] = render_ms
    profile_stats["updated_at"] = time.time()
    render_stats[width_profile] = profile_stats
    if len(render_stats) > MAX_RENDER_TELEMETRY_PROFILES:
        stale_profiles = sorted(
            render_stats.items(),
            key=lambda item: float(item[1].get("updated_at", 0.0)),
        )
        overflow = len(render_stats) - MAX_RENDER_TELEMETRY_PROFILES
        for profile_name, _stats in stale_profiles[:overflow]:
            render_stats.pop(profile_name, None)
    st.session_state["streamlit_render_stats"] = render_stats
    table_state = st.session_state.get("streamlit_table_state", {})
    _persist_streamlit_state(
        width_profile=str(table_state.get("width_profile", "Padrao (1600)")),
        width_profile_by_bucket=_normalize_width_profile_memory(
            table_state.get("width_profile_by_bucket", {})
        ),
        streamlit_render_stats=render_stats,
    )


def _resolve_width_bucket(width_px: int) -> str:
    if width_px < 1280:
        return "xs"
    if width_px < 1600:
        return "sm"
    if width_px < 2000:
        return "md"
    if width_px < 2400:
        return "lg"
    return "xl"


def _normalize_width_profile_memory(memory_raw: Any) -> dict[str, str]:
    if not isinstance(memory_raw, dict):
        return {}
    valid_buckets = {"xs", "sm", "md", "lg", "xl"}
    normalized: dict[str, str] = {}
    for key, value in memory_raw.items():
        key_text = str(key)
        value_text = str(value)
        if key_text in valid_buckets and value_text in WIDTH_PROFILE_OPTIONS:
            normalized[key_text] = value_text
    return normalized


def _normalize_streamlit_theme_name(raw_theme: Any) -> str:
    theme_name = str(raw_theme or DEFAULT_STREAMLIT_THEME)
    if theme_name not in STREAMLIT_THEME_PALETTES:
        return DEFAULT_STREAMLIT_THEME
    return theme_name


def _build_streamlit_theme_css(theme_name: str) -> str:
    theme = STREAMLIT_THEME_PALETTES[_normalize_streamlit_theme_name(theme_name)]
    input_bg = theme.get("input_bg", theme["panel"])
    input_ink = theme.get("input_ink", theme["ink"])
    return (
        "<style>"
        ":root {"
        f"--ssa-bg:{theme['bg']};"
        f"--ssa-panel:{theme['panel']};"
        f"--ssa-ink:{theme['ink']};"
        f"--ssa-muted:{theme['muted']};"
        f"--ssa-accent:{theme['accent']};"
        f"--ssa-accent-soft:{theme['accent_soft']};"
        f"--ssa-border:{theme['border']};"
        f"--ssa-metric-bg:{theme['metric_bg']};"
        f"--ssa-input-bg:{input_bg};"
        f"--ssa-input-ink:{input_ink};"
        "}"
        ".stApp{background:var(--ssa-bg);color:var(--ssa-ink);}"
        "section[data-testid='stSidebar']{background:var(--ssa-panel)!important;color:var(--ssa-ink)!important;}"
        ".block-container{padding-top:0.6rem;padding-bottom:0.5rem;}"
        "h1,h2,h3{color:var(--ssa-ink);}"
        ".section-label{font-size:0.84rem;color:var(--ssa-accent);margin-bottom:0.35rem;}"
        ".stButton button{width:100%;border:1px solid var(--ssa-border);}"
        "div[data-baseweb='tab-list']{position:sticky;top:0;z-index:995;background:var(--ssa-bg);"
        "border:1px solid var(--ssa-border);border-radius:0.55rem;padding:0.12rem;}"
        "div[data-baseweb='tab']{color:var(--ssa-ink)!important;}"
        "button[role='tab'][aria-selected='true']{background:var(--ssa-accent-soft)!important;"
        "border-color:var(--ssa-accent)!important;}"
        "div[data-testid='stMetric']{background:var(--ssa-metric-bg);border:1px solid var(--ssa-border);"
        "padding:0.5rem 0.65rem;border-radius:0.5rem;}"
        "div[data-testid='stMetricLabel'] *{color:var(--ssa-muted)!important;}"
        "div[data-testid='stMetricValue'] *{color:var(--ssa-ink)!important;}"
        "section[data-testid='stSidebar'] div[data-testid='stMetricLabel']{font-size:0.74rem;}"
        "section[data-testid='stSidebar'] div[data-testid='stMetricValue']{font-size:1.35rem;}"
        "div[data-testid='stDataFrame']{border:1px solid var(--ssa-border);border-radius:0.55rem;"
        "background:var(--ssa-panel);}"
        "div[data-testid='stDataFrame'] div[role='grid']{background:var(--ssa-panel)!important;"
        "color:var(--ssa-ink)!important;}"
        "div[data-testid='stDataFrame'] div[role='columnheader']{background:var(--ssa-accent-soft)!important;"
        "color:var(--ssa-ink)!important;}"
        "div[data-testid='stDataFrame'] div[role='gridcell']{color:var(--ssa-ink)!important;}"
        "div[data-testid='stForm']{background:var(--ssa-panel);border:1px solid var(--ssa-border);"
        "padding:0.6rem 0.75rem;border-radius:0.55rem;}"
        "div[data-testid='stHorizontalBlock'] > div{gap:0.45rem;}"
        "div[data-testid='stSelectbox'] > div,div[data-testid='stNumberInput'] > div,"
        "div[data-testid='stTextInput'] > div,div[data-testid='stDateInput'] > div,"
        "div[data-testid='stMultiSelect'] > div{"
        "border-color:var(--ssa-border);background:var(--ssa-input-bg)!important;color:var(--ssa-input-ink)!important;}"
        "div[data-testid='stTextInput'] input,div[data-testid='stDateInput'] input,div[data-testid='stNumberInput'] input{"
        "color:var(--ssa-input-ink)!important;}"
        "div[data-baseweb='select'] *{color:var(--ssa-input-ink)!important;}"
        "label,p,span,small{color:var(--ssa-ink);}"
        ".stCaption{color:var(--ssa-muted)!important;}"
        ".ssa-chip{display:inline-block;border:1px solid var(--ssa-border);"
        "background:var(--ssa-accent-soft);color:var(--ssa-ink);"
        "padding:0.1rem 0.45rem;border-radius:999px;margin-right:0.35rem;font-size:0.78rem;}"
        "</style>"
    )


def _normalize_render_stats(raw_stats: Any) -> dict[str, dict[str, Any]]:
    if not isinstance(raw_stats, dict):
        return {}
    normalized: dict[str, dict[str, Any]] = {}
    for raw_profile, raw_values in raw_stats.items():
        profile = str(raw_profile)
        if profile not in WIDTH_PROFILE_OPTIONS:
            continue
        if not isinstance(raw_values, dict):
            continue
        try:
            count = int(raw_values.get("count", 0))
            total_ms = float(raw_values.get("total_ms", 0.0))
            last_ms = float(raw_values.get("last_ms", 0.0))
            updated_at = float(raw_values.get("updated_at", 0.0))
        except (TypeError, ValueError):
            continue
        if count <= 0 or total_ms < 0.0 or last_ms < 0.0:
            continue
        normalized[profile] = {
            "count": count,
            "total_ms": total_ms,
            "last_ms": last_ms,
            "updated_at": updated_at,
        }
    if len(normalized) <= MAX_RENDER_TELEMETRY_PROFILES:
        return normalized
    stale_profiles = sorted(
        normalized.items(),
        key=lambda item: float(item[1].get("updated_at", 0.0)),
    )
    overflow = len(normalized) - MAX_RENDER_TELEMETRY_PROFILES
    for profile_name, _stats in stale_profiles[:overflow]:
        normalized.pop(profile_name, None)
    return normalized


def _resolve_streamlit_ui_state_path() -> Path:
    cfg_dir_raw = os.environ.get("SSA_CONFIG_DIR", "config")
    try:
        cfg_dir = ensure_path_is_allowed(
            cfg_dir_raw,
            purpose="SSA_CONFIG_DIR",
            base=project_root,
            expect_directory=None,
        )
    except PathSafetyError as exc:
        logger.warning(
            "SSA_CONFIG_DIR invalido para streamlit state (%s). Usando config padrao.",
            exc,
        )
        cfg_dir = project_root / "config"

    file_name = os.environ.get(
        "SSA_STREAMLIT_UI_STATE_FILE", STREAMLIT_UI_STATE_FILE_DEFAULT
    ).strip()
    if not file_name:
        file_name = STREAMLIT_UI_STATE_FILE_DEFAULT
    candidate = Path(file_name)
    if not candidate.is_absolute():
        candidate = cfg_dir / candidate
    return ensure_path_is_allowed(
        candidate,
        purpose="streamlit_ui_state_file",
        base=project_root,
        expect_directory=False,
    )


def _load_persisted_streamlit_state() -> dict[str, Any]:
    try:
        state_path = _resolve_streamlit_ui_state_path()
    except PathSafetyError as exc:
        logger.warning("Nao foi possivel resolver streamlit_ui_state_file: %s", exc)
        return {}
    if not state_path.exists():
        return {}
    try:
        with open(state_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except (OSError, json.JSONDecodeError, TypeError, ValueError) as exc:
        logger.warning(
            "Falha ao ler estado persistido do Streamlit (%s): %s", state_path, exc
        )
        return {}
    if not isinstance(payload, dict):
        return {}

    width_profile = str(payload.get("width_profile", "Padrao (1600)"))
    if width_profile not in WIDTH_PROFILE_OPTIONS:
        width_profile = "Padrao (1600)"
    return {
        "theme_name": _normalize_streamlit_theme_name(
            payload.get("theme_name", DEFAULT_STREAMLIT_THEME)
        ),
        "width_profile": width_profile,
        "width_profile_by_bucket": _normalize_width_profile_memory(
            payload.get("width_profile_by_bucket", {})
        ),
        "streamlit_render_stats": _normalize_render_stats(
            payload.get("streamlit_render_stats", {})
        ),
    }


def _persist_streamlit_state(
    *,
    theme_name: str | None = None,
    width_profile: str,
    width_profile_by_bucket: dict[str, str],
    streamlit_render_stats: dict[str, Any],
) -> None:
    if width_profile not in WIDTH_PROFILE_OPTIONS:
        width_profile = "Padrao (1600)"
    payload = {
        "theme_name": _normalize_streamlit_theme_name(
            theme_name
            if theme_name is not None
            else (
                st.session_state.get("streamlit_theme_name", DEFAULT_STREAMLIT_THEME)
                if hasattr(st, "session_state") and st.session_state is not None
                else DEFAULT_STREAMLIT_THEME
            )
        ),
        "width_profile": width_profile,
        "width_profile_by_bucket": _normalize_width_profile_memory(
            width_profile_by_bucket
        ),
        "streamlit_render_stats": _normalize_render_stats(streamlit_render_stats),
        "updated_at": time.time(),
    }
    try:
        state_path = _resolve_streamlit_ui_state_path()
        state_path.parent.mkdir(parents=True, exist_ok=True)
        config_manager.atomic_write_json_file(
            str(state_path),
            payload,
            indent=2,
            ensure_ascii=False,
        )
    except Exception as exc:
        logger.warning("Falha ao persistir estado opcional do Streamlit: %s", exc)


def _resolve_width_profile_for_bucket(table_state: dict[str, Any]) -> tuple[str, str]:
    fallback_profile = str(table_state.get("width_profile", "Padrao (1600)"))
    if fallback_profile not in WIDTH_PROFILE_OPTIONS:
        fallback_profile = "Padrao (1600)"

    width_px = WIDTH_PROFILE_PIXELS.get(fallback_profile, 1600)
    if hasattr(st, "session_state") and st.session_state is not None:
        viewport_hint_raw = st.session_state.get("streamlit_viewport_width_px")
        if viewport_hint_raw is not None:
            try:
                width_px = int(viewport_hint_raw)
                if width_px <= 0:
                    width_px = WIDTH_PROFILE_PIXELS.get(fallback_profile, 1600)
            except (TypeError, ValueError):
                width_px = WIDTH_PROFILE_PIXELS.get(fallback_profile, 1600)

    width_bucket = _resolve_width_bucket(width_px)
    memory = _normalize_width_profile_memory(
        table_state.get("width_profile_by_bucket", {})
    )
    selected_profile = str(memory.get(width_bucket, fallback_profile))
    if selected_profile not in WIDTH_PROFILE_OPTIONS:
        selected_profile = fallback_profile
    return selected_profile, width_bucket


def _remember_width_profile_for_bucket(
    table_state: dict[str, Any],
    width_bucket: str,
    width_profile: str,
) -> None:
    memory = _normalize_width_profile_memory(
        table_state.get("width_profile_by_bucket", {})
    )
    if width_profile in WIDTH_PROFILE_OPTIONS:
        memory[str(width_bucket)] = width_profile
    table_state["width_profile_by_bucket"] = memory


def _api_snapshot_available(consult_api: bool, recent_df: pd.DataFrame | None) -> bool:
    return bool(consult_api and recent_df is not None and not recent_df.empty)


def _clear_recent_api_snapshot() -> None:
    if not hasattr(st, "session_state") or st.session_state is None:
        return
    st.session_state["recent_api_df"] = None


def _build_table_caption(
    compact_mode: bool,
    page_number: int,
    total_pages: int,
    page_len: int,
    filtered_len: int,
    render_ms: float,
) -> str:
    if compact_mode:
        return (
            f"Pag {page_number}/{total_pages} | linhas: {page_len}/{filtered_len} | "
            f"render: {render_ms:.1f} ms"
        )
    return (
        f"Exibindo pagina {page_number}/{total_pages} | "
        f"linhas nesta pagina: {page_len} | linhas filtradas: {filtered_len}"
    )


def _format_render_stats_line(profile: str, profile_stats: dict[str, Any]) -> str:
    avg_ms = float(profile_stats["total_ms"]) / max(1, int(profile_stats["count"]))
    return (
        f"Render tabela ({profile}): "
        f"ultimo {float(profile_stats['last_ms']):.1f} ms | media {avg_ms:.1f} ms"
    )


def apply_cli_filters(df: pd.DataFrame, search_text: str) -> pd.DataFrame:
    """Aplica filtros CLI com fallback para caso sem cache."""
    if not search_text.strip():
        return df
    raw_terms = [term.strip() for term in search_text.split(",") if term.strip()]
    parsed = parse_search_terms(raw_terms)
    return filter_dataframe(df, parsed)


def _normalize_date_boundary(raw_value: Any) -> pd.Timestamp | None:
    parsed_text = parse_any_date(raw_value)
    if not parsed_text:
        return None
    parsed = pd.to_datetime(parsed_text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed


def _parse_date_input(raw_value: Any) -> date | None:
    parsed_text = parse_any_date(raw_value)
    if not parsed_text:
        return None
    parsed = pd.to_datetime(parsed_text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _format_date_input(raw_value: Any) -> str:
    if raw_value in (None, ""):
        return ""
    parsed = pd.to_datetime(raw_value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return str(parsed.date())


def _date_to_yearweek(value: datetime | pd.Timestamp) -> int | None:
    try:
        timestamp = pd.Timestamp(value)
        if pd.isna(timestamp):
            return None
        safe_timestamp = cast(pd.Timestamp, timestamp)
        iso_value = safe_timestamp.isocalendar()
    except Exception:
        return None
    year_part = getattr(iso_value, "year", None)
    week_part = getattr(iso_value, "week", None)
    if year_part is not None and week_part is not None:
        try:
            return int(int(year_part) * 100 + int(week_part))
        except Exception:
            return None
    if isinstance(iso_value, (tuple, list)) and len(iso_value) >= 2:
        try:
            return int(int(iso_value[0]) * 100 + int(iso_value[1]))
        except Exception:
            return None
    return None


def _compute_sidebar_weekly_kpis(
    df: pd.DataFrame,
    reference_dt: Optional[datetime | pd.Timestamp] = None,
) -> dict[str, int]:
    empty = {
        "executadas_semana_atual": 0,
        "executadas_semana_anterior": 0,
        "emitidas_semana_atual": 0,
        "emitidas_semana_anterior": 0,
    }
    if df.empty:
        return empty

    now_dt = reference_dt or datetime.now()
    current_week = _date_to_yearweek(now_dt)
    previous_week = _date_to_yearweek(now_dt - timedelta(days=7))
    if current_week is None or previous_week is None:
        return empty

    exec_week = pd.Series(pd.array([pd.NA] * len(df), dtype="Int64"), index=df.index)
    if "data_execucao" in df.columns:
        exec_week = _compute_yearweek_from_date_series(df, "data_execucao")
    elif "semana_executada" in df.columns:
        exec_week = _compute_yearweek_from_week_series(df, "semana_executada")

    emit_week = pd.Series(pd.array([pd.NA] * len(df), dtype="Int64"), index=df.index)
    if "data_emissao" in df.columns:
        emit_week = _compute_yearweek_from_date_series(df, "data_emissao")
    elif "data_cadastro" in df.columns:
        emit_week = _compute_yearweek_from_date_series(df, "data_cadastro")

    return {
        "executadas_semana_atual": int(exec_week.eq(current_week).fillna(False).sum()),
        "executadas_semana_anterior": int(
            exec_week.eq(previous_week).fillna(False).sum()
        ),
        "emitidas_semana_atual": int(emit_week.eq(current_week).fillna(False).sum()),
        "emitidas_semana_anterior": int(
            emit_week.eq(previous_week).fillna(False).sum()
        ),
    }


def _compute_sidebar_weekly_kpis_cached(
    df: pd.DataFrame, reference_day: str
) -> dict[str, int]:
    try:
        reference_dt = datetime.strptime(reference_day, "%Y-%m-%d")
    except ValueError:
        reference_dt = datetime.now()
    return _compute_sidebar_weekly_kpis(df, reference_dt=reference_dt)


if hasattr(st, "cache_data") and callable(getattr(st, "cache_data")):
    _compute_sidebar_weekly_kpis_cached = st.cache_data(show_spinner=False)(
        _compute_sidebar_weekly_kpis_cached
    )


def _compute_year_from_date_series(df: pd.DataFrame, col_name: str) -> pd.Series:
    if col_name not in df.columns:
        return pd.Series(pd.array([pd.NA] * len(df), dtype="Int64"), index=df.index)
    parsed = pd.to_datetime(df[col_name], errors="coerce")
    years = parsed.dt.year
    return years.astype("Int64")


def _compute_year_from_week_series(df: pd.DataFrame, col_name: str) -> pd.Series:
    if col_name not in df.columns:
        return pd.Series(pd.array([pd.NA] * len(df), dtype="Int64"), index=df.index)
    nums = pd.to_numeric(df[col_name], errors="coerce").astype("Int64")
    years = nums // 100
    return years.astype("Int64")


def _compute_yearweek_from_date_series(df: pd.DataFrame, col_name: str) -> pd.Series:
    if col_name not in df.columns:
        return pd.Series(pd.array([pd.NA] * len(df), dtype="Int64"), index=df.index)
    parsed = pd.to_datetime(df[col_name], errors="coerce")
    iso = parsed.dt.isocalendar()
    yearweek = (iso["year"] * 100) + iso["week"]
    return yearweek.astype("Int64")


def _compute_yearweek_from_week_series(df: pd.DataFrame, col_name: str) -> pd.Series:
    if col_name not in df.columns:
        return pd.Series(pd.array([pd.NA] * len(df), dtype="Int64"), index=df.index)
    return pd.to_numeric(df[col_name], errors="coerce").astype("Int64")


def _compute_derivada_flags(df: pd.DataFrame) -> tuple[pd.Series, pd.Series, pd.Series]:
    if "numero_ssa" not in df.columns or "derivada_de" not in df.columns:
        false_series = pd.Series(False, index=df.index)
        zero_series = pd.Series(0, index=df.index, dtype="Int64")
        return false_series, false_series, zero_series

    numero = df["numero_ssa"].fillna("").astype(str).str.strip()
    derivada = df["derivada_de"].fillna("").astype(str).str.strip()
    invalid = {"", "nan", "none", "null"}
    has_derivada = ~derivada.str.casefold().isin(invalid)
    valid_parent = derivada[has_derivada]
    counts = valid_parent.value_counts()
    qtd_derivadas = numero.map(counts).fillna(0).astype("Int64")
    has_children = qtd_derivadas.gt(0)
    return has_derivada, has_children, qtd_derivadas


def _build_derivada_context_map(df: pd.DataFrame) -> pd.DataFrame:
    if "numero_ssa" not in df.columns:
        return pd.DataFrame(
            columns=[
                "numero_ssa",
                "_tem_derivada_ctx",
                "_tem_derivadas_ctx",
                "_qtd_derivadas_ctx",
            ]
        )
    has_derivada, has_children, qtd_derivadas = _compute_derivada_flags(df)
    numero_series = df["numero_ssa"].fillna("").astype(str).str.strip()
    context_map = pd.DataFrame(
        {
            "numero_ssa": numero_series,
            "_tem_derivada_ctx": has_derivada.map({True: "Sim", False: "Nao"}),
            "_tem_derivadas_ctx": has_children.map({True: "Sim", False: "Nao"}),
            "_qtd_derivadas_ctx": qtd_derivadas.fillna(0).astype(int),
        }
    ).drop_duplicates(subset=["numero_ssa"], keep="first")
    return context_map


def _attach_derivada_context(
    df: pd.DataFrame, base_context_map: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    if df.empty:
        return df
    if "numero_ssa" not in df.columns:
        return df
    context_map = (
        base_context_map if isinstance(base_context_map, pd.DataFrame) else None
    )
    if context_map is None:
        if "derivada_de" not in df.columns:
            return df
        context_map = _build_derivada_context_map(df)
    df.attrs["_derivada_context_map"] = context_map
    return df


def _build_active_summary(
    *,
    search_terms: str,
    situacao_filter: list[Any],
    executor_filter: list[Any],
    emissor_filter: list[Any],
    advanced_filters: dict[str, Any],
    consult_api: bool,
) -> list[str]:
    summary: list[str] = []

    def _append_values(label: str, values: list[Any]) -> None:
        if values:
            summary.append(f"{label}: " + ", ".join(str(value) for value in values))

    if search_terms.strip():
        summary.append(f"Busca: {search_terms.strip()}")
    if situacao_filter:
        summary.append(
            "Situacao: " + ", ".join(str(value) for value in situacao_filter)
        )
    if executor_filter:
        summary.append(
            "Executor: " + ", ".join(str(value) for value in executor_filter)
        )
    if emissor_filter:
        summary.append("Emissor: " + ", ".join(str(value) for value in emissor_filter))
    _append_values("Executor(resp)", list(advanced_filters.get("executor_resp") or []))
    _append_values(
        "Executor(resp)(!)", list(advanced_filters.get("executor_resp_exclude") or [])
    )
    _append_values("Estado+", list(advanced_filters.get("estado") or []))
    _append_values("Estado+(!)", list(advanced_filters.get("estado_exclude") or []))
    _append_values("Ano emissao", list(advanced_filters.get("ano_emissao") or []))
    _append_values(
        "Ano emissao(!)", list(advanced_filters.get("ano_emissao_exclude") or [])
    )
    _append_values("Ano execucao", list(advanced_filters.get("ano_execucao") or []))
    _append_values(
        "Ano execucao(!)", list(advanced_filters.get("ano_execucao_exclude") or [])
    )
    _append_values(
        "AnoSemana emissao", list(advanced_filters.get("ano_semana_emissao") or [])
    )
    _append_values(
        "AnoSemana emissao(!)",
        list(advanced_filters.get("ano_semana_emissao_exclude") or []),
    )
    _append_values(
        "AnoSemana execucao", list(advanced_filters.get("ano_semana_execucao") or [])
    )
    _append_values(
        "AnoSemana execucao(!)",
        list(advanced_filters.get("ano_semana_execucao_exclude") or []),
    )
    _append_values(
        "Reprogramacoes", list(advanced_filters.get("num_reprogramacoes") or [])
    )
    _append_values(
        "Reprogramacoes(!)",
        list(advanced_filters.get("num_reprogramacoes_exclude") or []),
    )
    if advanced_filters["tem_derivada"] != "todos":
        summary.append(f"Tem derivada: {advanced_filters['tem_derivada']}")
    if advanced_filters["tem_derivadas"] != "todos":
        summary.append(f"Tem derivadas: {advanced_filters['tem_derivadas']}")
    if advanced_filters["data_emissao_inicio"] or advanced_filters["data_emissao_fim"]:
        summary.append(
            "Data emissao: "
            + str(advanced_filters["data_emissao_inicio"] or "-")
            + " ate "
            + str(advanced_filters["data_emissao_fim"] or "-")
        )
    if (
        advanced_filters["data_execucao_inicio"]
        or advanced_filters["data_execucao_fim"]
    ):
        summary.append(
            "Data execucao: "
            + str(advanced_filters["data_execucao_inicio"] or "-")
            + " ate "
            + str(advanced_filters["data_execucao_fim"] or "-")
        )
    if consult_api:
        summary.append("API: manual")
    return summary


def _apply_include_exclude_filter(
    filtered_df: pd.DataFrame,
    *,
    column_name: str,
    include_values: list[Any],
    exclude_values: list[Any],
) -> pd.DataFrame:
    if column_name not in filtered_df.columns:
        return filtered_df
    if include_values:
        filtered_df = filtered_df[filtered_df[column_name].isin(include_values)]
    if exclude_values:
        filtered_df = filtered_df[~filtered_df[column_name].isin(exclude_values)]
    return filtered_df


def _apply_numeric_series_include_exclude_filter(
    filtered_df: pd.DataFrame,
    *,
    series: pd.Series,
    include_values: list[int],
    exclude_values: list[int],
) -> pd.DataFrame:
    if include_values:
        filtered_df = filtered_df[series.isin(set(include_values))]
        series = series[filtered_df.index]
    if exclude_values:
        filtered_df = filtered_df[~series.isin(set(exclude_values))]
    return filtered_df


def _apply_datetime_manual_boundary_filter(
    filtered_df: pd.DataFrame,
    *,
    column_name: str,
    start_raw: Any,
    end_raw: Any,
) -> pd.DataFrame:
    if column_name not in filtered_df.columns:
        return filtered_df
    start_conditions = _parse_manual_boundary_conditions(start_raw, is_start=True)
    end_conditions = _parse_manual_boundary_conditions(end_raw, is_start=False)
    if not start_conditions and not end_conditions:
        return filtered_df
    parsed = pd.to_datetime(filtered_df[column_name], errors="coerce")
    mask = pd.Series(True, index=filtered_df.index)
    for op, boundary in start_conditions + end_conditions:
        if op == "ge":
            current_mask = parsed.ge(boundary).fillna(False)
        elif op == "le":
            current_mask = parsed.le(boundary).fillna(False)
        elif op == "lt":
            current_mask = parsed.lt(boundary).fillna(False)
        elif op == "ne":
            current_mask = parsed.ne(boundary).fillna(False)
        else:
            current_mask = parsed.gt(boundary).fillna(False)
        mask &= current_mask
    return filtered_df[mask]


def _apply_yearweek_manual_boundary_filter(
    filtered_df: pd.DataFrame,
    *,
    week_column: str,
    start_raw: Any,
    end_raw: Any,
) -> pd.DataFrame:
    if week_column not in filtered_df.columns:
        return filtered_df
    start_conditions = _parse_manual_boundary_conditions(start_raw, is_start=True)
    end_conditions = _parse_manual_boundary_conditions(end_raw, is_start=False)
    if not start_conditions and not end_conditions:
        return filtered_df
    week_series = _compute_yearweek_from_week_series(filtered_df, week_column)
    mask = pd.Series(True, index=filtered_df.index)
    for op, boundary in start_conditions + end_conditions:
        year_week = _date_to_yearweek(boundary)
        if year_week is None:
            continue
        if op == "ge":
            current_mask = week_series.ge(year_week).fillna(False)
        elif op == "le":
            current_mask = week_series.le(year_week).fillna(False)
        elif op == "lt":
            current_mask = week_series.lt(year_week).fillna(False)
        elif op == "ne":
            current_mask = week_series.ne(year_week).fillna(False)
        else:
            current_mask = week_series.gt(year_week).fillna(False)
        mask &= current_mask
    return filtered_df[mask]


def _apply_advanced_streamlit_filters(
    filtered_df: pd.DataFrame, advanced_filters: Optional[dict[str, Any]]
) -> pd.DataFrame:
    adv = advanced_filters or {}
    if not adv:
        return filtered_df

    filtered_df = _apply_include_exclude_filter(
        filtered_df,
        column_name="responsavel_execucao",
        include_values=list(adv.get("executor_resp") or []),
        exclude_values=list(adv.get("executor_resp_exclude") or []),
    )
    filtered_df = _apply_include_exclude_filter(
        filtered_df,
        column_name="situacao",
        include_values=list(adv.get("estado") or []),
        exclude_values=list(adv.get("estado_exclude") or []),
    )

    ano_emissao_values = [
        int(v) for v in (adv.get("ano_emissao") or []) if str(v).isdigit()
    ]
    ano_emissao_exclude_values = [
        int(v) for v in (adv.get("ano_emissao_exclude") or []) if str(v).isdigit()
    ]
    if ano_emissao_values or ano_emissao_exclude_values:
        ano_emissao_series = _compute_year_from_date_series(
            filtered_df, "data_cadastro"
        )
        filtered_df = _apply_numeric_series_include_exclude_filter(
            filtered_df,
            series=ano_emissao_series,
            include_values=ano_emissao_values,
            exclude_values=ano_emissao_exclude_values,
        )

    ano_execucao_values = [
        int(v) for v in (adv.get("ano_execucao") or []) if str(v).isdigit()
    ]
    ano_execucao_exclude_values = [
        int(v) for v in (adv.get("ano_execucao_exclude") or []) if str(v).isdigit()
    ]
    if ano_execucao_values or ano_execucao_exclude_values:
        ano_execucao_series = _compute_year_from_week_series(
            filtered_df, "semana_executada"
        )
        filtered_df = _apply_numeric_series_include_exclude_filter(
            filtered_df,
            series=ano_execucao_series,
            include_values=ano_execucao_values,
            exclude_values=ano_execucao_exclude_values,
        )

    ano_semana_emissao_values = [
        int(v) for v in (adv.get("ano_semana_emissao") or []) if str(v).isdigit()
    ]
    ano_semana_emissao_exclude_values = [
        int(v)
        for v in (adv.get("ano_semana_emissao_exclude") or [])
        if str(v).isdigit()
    ]
    if ano_semana_emissao_values or ano_semana_emissao_exclude_values:
        ano_semana_emissao_series = _compute_yearweek_from_date_series(
            filtered_df, "data_cadastro"
        )
        filtered_df = _apply_numeric_series_include_exclude_filter(
            filtered_df,
            series=ano_semana_emissao_series,
            include_values=ano_semana_emissao_values,
            exclude_values=ano_semana_emissao_exclude_values,
        )

    ano_semana_execucao_values = [
        int(v) for v in (adv.get("ano_semana_execucao") or []) if str(v).isdigit()
    ]
    ano_semana_execucao_exclude_values = [
        int(v)
        for v in (adv.get("ano_semana_execucao_exclude") or [])
        if str(v).isdigit()
    ]
    if ano_semana_execucao_values or ano_semana_execucao_exclude_values:
        ano_semana_execucao_series = _compute_yearweek_from_week_series(
            filtered_df, "semana_executada"
        )
        filtered_df = _apply_numeric_series_include_exclude_filter(
            filtered_df,
            series=ano_semana_execucao_series,
            include_values=ano_semana_execucao_values,
            exclude_values=ano_semana_execucao_exclude_values,
        )

    reprog_values = [
        int(v) for v in (adv.get("num_reprogramacoes") or []) if str(v).isdigit()
    ]
    reprog_exclude_values = [
        int(v)
        for v in (adv.get("num_reprogramacoes_exclude") or [])
        if str(v).isdigit()
    ]
    if (
        reprog_values or reprog_exclude_values
    ) and "num_reprogramacoes" in filtered_df.columns:
        nums = pd.to_numeric(filtered_df["num_reprogramacoes"], errors="coerce").astype(
            "Int64"
        )
        filtered_df = _apply_numeric_series_include_exclude_filter(
            filtered_df,
            series=nums,
            include_values=reprog_values,
            exclude_values=reprog_exclude_values,
        )

    filtered_df = _apply_datetime_manual_boundary_filter(
        filtered_df,
        column_name="data_cadastro",
        start_raw=adv.get("data_emissao_inicio"),
        end_raw=adv.get("data_emissao_fim"),
    )
    filtered_df = _apply_yearweek_manual_boundary_filter(
        filtered_df,
        week_column="semana_executada",
        start_raw=adv.get("data_execucao_inicio"),
        end_raw=adv.get("data_execucao_fim"),
    )

    tem_derivada_mode = str(adv.get("tem_derivada") or "todos").strip().lower()
    tem_derivadas_mode = str(adv.get("tem_derivadas") or "todos").strip().lower()
    if tem_derivada_mode != "todos" or tem_derivadas_mode != "todos":
        has_derivada, has_children, _ = _compute_derivada_flags(filtered_df)
        if tem_derivada_mode == "sim":
            filtered_df = filtered_df[has_derivada]
            has_children = has_children[has_derivada]
        elif tem_derivada_mode == "nao":
            filtered_df = filtered_df[~has_derivada]
            has_children = has_children[~has_derivada]
        if tem_derivadas_mode == "sim":
            filtered_df = filtered_df[has_children]
        elif tem_derivadas_mode == "nao":
            filtered_df = filtered_df[~has_children]

    return filtered_df


def _normalize_cache_value(value: Any) -> Any:
    if isinstance(value, dict):
        return tuple(
            (str(k), _normalize_cache_value(v))
            for k, v in sorted(value.items(), key=lambda item: str(item[0]))
        )
    if isinstance(value, list):
        normalized_items = [_normalize_cache_value(v) for v in value]
        return tuple(sorted(normalized_items, key=lambda v: str(v)))
    return str(value)


def apply_all_filters_cached(
    df: pd.DataFrame,
    search_terms: str,
    situacoes: list,
    executores: list,
    emissores: list,
    *,
    advanced_filters: Optional[dict[str, Any]] = None,
    base_derivada_context_map: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Aplica todos os filtros com cache inteligente."""
    df_token = _compute_df_cache_token(df)
    # Verifica cache primeiro
    cached_result = filter_cache.get(
        df.shape,
        search_terms,
        situacoes,
        executores,
        emissores,
        advanced_filters=advanced_filters,
        df_token=df_token,
    )
    if cached_result is not None:
        return cached_result

    # Cache miss - aplica filtros
    start_time = time.time()

    # Filtro de busca textual
    filtered_df = apply_cli_filters(df, search_terms)

    # Filtros de selecao multipla
    if situacoes and "situacao" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["situacao"].isin(situacoes)]
    if executores and "setor_executor" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["setor_executor"].isin(executores)]
    if emissores and "setor_emissor" in filtered_df.columns:
        filtered_df = filtered_df[filtered_df["setor_emissor"].isin(emissores)]

    filtered_df = _apply_advanced_streamlit_filters(filtered_df, advanced_filters)

    # Reset index
    filtered_df = filtered_df.reset_index(drop=True)
    filtered_df = _attach_derivada_context(
        filtered_df, base_context_map=base_derivada_context_map
    )

    # Armazena no cache
    filter_cache.put(
        df.shape,
        search_terms,
        situacoes,
        executores,
        emissores,
        filtered_df,
        advanced_filters=advanced_filters,
        df_token=df_token,
    )

    # Log performance se demorou mais que 100ms
    elapsed = time.time() - start_time
    if elapsed > 0.1:
        logger.info("Filtro streamlit executado em %.2fs (cache miss)", elapsed)

    return filtered_df


def ensure_arrow_compatible(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza colunas para evitar falhas do Streamlit/Arrow.

    Tenta preservar tipos numericos quando possivel, convertendo apenas
    colunas verdadeiramente mistas para string.
    """
    safe = df.copy()  # Cheap with copy-on-write
    conversions_made = []

    for col in safe.columns:
        series = safe[col]
        dtype_str = str(series.dtype)

        # Skip modern nullable dtypes (already Arrow-compatible)
        if dtype_str.startswith("string") or dtype_str in [
            "Int64",
            "Int32",
            "Float64",
            "Float32",
        ]:
            continue

        if series.dtype == "object":
            non_null = series.dropna()
            if not non_null.empty:
                sample_types = {type(x) for x in non_null.head(20)}

                # Check if truly mixed types
                if len(sample_types) > 1:
                    # Sample more values to determine majority type
                    sample_size = min(100, len(non_null))
                    type_counts: dict[type[Any], int] = {}
                    for val in non_null.head(sample_size):
                        t = type(val)
                        type_counts[t] = type_counts.get(t, 0) + 1

                    majority_type = max(type_counts, key=type_counts.__getitem__)

                    # If mostly numeric, try to preserve as numeric
                    if majority_type in (int, float):
                        try:
                            safe[col] = pd.to_numeric(series, errors="coerce")
                            conversions_made.append(f"{col}: mixed->numeric")
                            continue
                        except Exception:
                            pass  # Fall through to string conversion

                    # Otherwise convert to string
                    safe[col] = series.astype(str)
                    conversions_made.append(f"{col}: mixed->string")
                    continue

                # Handle list/dict types
                sample = non_null.iloc[0]
                if isinstance(sample, (list, dict)):
                    safe[col] = series.apply(
                        lambda x: json.dumps(x, ensure_ascii=False)
                        if isinstance(x, (list, dict))
                        else x
                    )
                    conversions_made.append(f"{col}: list/dict->json")

        elif pd.api.types.is_integer_dtype(series.dtype):
            safe[col] = series.astype("Int64")

    if conversions_made:
        logger.debug(f"Arrow compatibility conversions: {', '.join(conversions_made)}")

    return safe


def _build_filter_options(df: pd.DataFrame) -> tuple[list[Any], list[Any], list[Any]]:
    situacoes = sorted(
        df.get("situacao", pd.Series(dtype=str)).dropna().unique().tolist(),
        key=lambda value: str(value),
    )
    executores = sorted(
        df.get("setor_executor", pd.Series(dtype=str)).dropna().unique().tolist(),
        key=lambda value: str(value),
    )
    emissores = sorted(
        df.get("setor_emissor", pd.Series(dtype=str)).dropna().unique().tolist(),
        key=lambda value: str(value),
    )
    return situacoes, executores, emissores


def _build_advanced_filter_options(df: pd.DataFrame) -> dict[str, list[Any]]:
    def _sorted_unique(col_name: str) -> list[Any]:
        values = df.get(col_name, pd.Series(dtype=str)).dropna().unique().tolist()
        return sorted(values, key=lambda value: str(value))

    ano_emissao = (
        _compute_year_from_date_series(df, "data_cadastro")
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    ano_execucao = (
        _compute_year_from_week_series(df, "semana_executada")
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    ano_semana_emissao = (
        _compute_yearweek_from_date_series(df, "data_cadastro")
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    ano_semana_execucao = (
        _compute_yearweek_from_week_series(df, "semana_executada")
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    reprog_values = (
        pd.to_numeric(
            df.get("num_reprogramacoes", pd.Series(dtype="Int64")), errors="coerce"
        )
        .dropna()
        .astype(int)
        .unique()
        .tolist()
    )
    return {
        "executor_resp": _sorted_unique("responsavel_execucao"),
        "estado": _sorted_unique("situacao"),
        "ano_emissao": sorted(ano_emissao, reverse=True),
        "ano_execucao": sorted(ano_execucao, reverse=True),
        "ano_semana_emissao": sorted(ano_semana_emissao, reverse=True),
        "ano_semana_execucao": sorted(ano_semana_execucao, reverse=True),
        "num_reprogramacoes": sorted(reprog_values),
    }


def _build_table_with_derivada_context(
    view_df: pd.DataFrame, filtered_df: pd.DataFrame
) -> pd.DataFrame:
    if view_df.empty:
        return view_df
    if "numero_ssa" not in filtered_df.columns:
        return view_df

    context_map = filtered_df.attrs.get("_derivada_context_map")
    if not isinstance(context_map, pd.DataFrame):
        if "derivada_de" not in filtered_df.columns:
            return view_df
        context_map = _build_derivada_context_map(filtered_df)

    if "numero_ssa" not in view_df.columns:
        return view_df

    merged = view_df.merge(context_map, on="numero_ssa", how="left")
    merged["_tem_derivada_ctx"] = merged["_tem_derivada_ctx"].fillna("Nao")
    merged["_tem_derivadas_ctx"] = merged["_tem_derivadas_ctx"].fillna("Nao")
    merged["_qtd_derivadas_ctx"] = (
        pd.to_numeric(
            merged["_qtd_derivadas_ctx"],
            errors="coerce",
        )
        .fillna(0)
        .astype(int)
    )
    return merged


def _paginate_dataframe(
    df: pd.DataFrame, page: int, page_size: int
) -> tuple[pd.DataFrame, int]:
    if page_size <= 0:
        raise ValueError("page_size must be greater than zero")
    total_pages = max(1, math.ceil(len(df) / page_size))
    current_page = min(max(page, 1), total_pages)
    start = (current_page - 1) * page_size
    end = start + page_size
    return df.iloc[start:end].reset_index(drop=True), total_pages


def _apply_large_page_guard(page_size: int, filtered_len: int) -> tuple[int, bool]:
    """Optional guard for very large pages. Disabled by default."""
    if page_size <= 0:
        return page_size, False
    if filtered_len <= 2000:
        return page_size, False
    if os.environ.get("SSA_STREAMLIT_LARGE_PAGE_GUARD", "0").strip() != "1":
        return page_size, False
    guarded = min(page_size, 500)
    return guarded, guarded != page_size


def _normalize_filter_selection(selected: list[Any], options: list[Any]) -> list[Any]:
    if not options:
        return []
    normalized = [value for value in selected if value in options]
    if len(normalized) == len(options):
        return []
    return normalized


def _split_manual_filter_tokens(raw_value: Any) -> tuple[list[str], list[str]]:
    include_tokens: list[str] = []
    exclude_tokens: list[str] = []
    raw_text = str(raw_value or "").strip()
    if not raw_text:
        return include_tokens, exclude_tokens
    for piece in raw_text.split(","):
        token = piece.strip()
        if not token:
            continue
        if token.startswith("!"):
            neg_token = token[1:].strip()
            if neg_token:
                exclude_tokens.append(neg_token)
            continue
        include_tokens.append(token)
    return include_tokens, exclude_tokens


def _resolve_include_exclude_values(
    selected: list[Any],
    options: list[Any],
    manual_text: Any,
) -> tuple[list[Any], list[Any]]:
    include_values = _normalize_filter_selection(selected, options)
    include_tokens, exclude_tokens = _split_manual_filter_tokens(manual_text)
    option_map = {str(value).strip().casefold(): value for value in options}

    for token in include_tokens:
        mapped = option_map.get(token.casefold())
        if mapped is not None and mapped not in include_values:
            include_values.append(mapped)

    exclude_values: list[Any] = []
    for token in exclude_tokens:
        mapped = option_map.get(token.casefold())
        if mapped is not None and mapped not in exclude_values:
            exclude_values.append(mapped)

    if include_values and exclude_values:
        include_values = [
            value for value in include_values if value not in exclude_values
        ]
    return include_values, exclude_values


def _parse_manual_boundary_conditions(
    raw_value: Any, *, is_start: bool
) -> list[tuple[str, pd.Timestamp]]:
    include_tokens, exclude_tokens = _split_manual_filter_tokens(raw_value)
    conditions: list[tuple[str, pd.Timestamp]] = []

    for token in include_tokens:
        parsed = _normalize_date_boundary(token)
        if parsed is None:
            continue
        conditions.append(("ge" if is_start else "le", parsed))
    for token in exclude_tokens:
        parsed = _normalize_date_boundary(token)
        if parsed is None:
            continue
        conditions.append(("ne", parsed))
    return conditions


def _default_visible_columns(columns: list[str]) -> list[str]:
    core = [
        "numero_ssa",
        "situacao",
        "descricao_ssa",
        "setor_executor",
        "setor_emissor",
        "data_cadastro",
        "prazo_limite",
    ]
    picked = [col for col in columns if col in core]
    return picked if picked else list(columns[:10])


def _build_column_presets(columns: list[str]) -> dict[str, list[str]]:
    defaults = _default_visible_columns(columns)
    return {
        "core": defaults,
        "all": list(columns),
    }


def _columns_with_data(df: pd.DataFrame, columns: list[str]) -> list[str]:
    valid: list[str] = []
    for col in columns:
        if col not in df.columns:
            continue
        series = df[col]
        if series.notna().any():
            valid.append(col)
    return valid


def _compute_table_render_height(page_len: int, configured_height: int) -> int:
    dynamic_height = 56 + (max(1, page_len) * 36)
    bounded = min(max(dynamic_height, 220), max(260, configured_height))
    return int(bounded)


def _resolve_situacao_quick_mode(
    situacoes: list[Any],
    manual_values: list[Any],
    mode: str,
) -> list[Any]:
    mode_key = str(mode).strip().lower()
    if mode_key == "todas":
        return list(situacoes)
    if mode_key == "executadas":
        picked = [value for value in situacoes if "EXECUT" in str(value).upper()]
        return picked if picked else list(situacoes)
    if mode_key == "abertas":
        picked = []
        for value in situacoes:
            text = str(value).upper()
            if "EXECUT" in text or "CONCL" in text or "FINAL" in text:
                continue
            picked.append(value)
        return picked if picked else list(situacoes)
    if mode_key == "nenhuma":
        return []
    normalized_manual = [value for value in manual_values if value in situacoes]
    return normalized_manual if normalized_manual else list(situacoes)


def _list_spreadsheet_files_count(docs_dir: str) -> int:
    try:
        docs_path = Path(docs_dir)
        if not docs_path.exists() or not docs_path.is_dir():
            return 0
        count = 0
        for file_path in docs_path.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in {
                ".xlsx",
                ".xls",
                ".csv",
            }:
                count += 1
        return count
    except Exception:
        return 0


def _is_real_streamlit_runtime() -> bool:
    # Evita executar UI completa fora do runtime do streamlit.
    try:
        runtime_module = getattr(st, "runtime", None)
        exists_fn = getattr(runtime_module, "exists", None)
        if callable(exists_fn):
            return bool(exists_fn())
        has_ui_api = callable(getattr(st, "set_page_config", None)) and callable(
            getattr(st, "columns", None)
        )
        has_state = hasattr(st, "session_state") and st.session_state is not None
        return bool(has_ui_api and has_state)
    except Exception:
        return False


REAL_RUNTIME = _is_real_streamlit_runtime()

if REAL_RUNTIME:
    persisted_boot_state = _load_persisted_streamlit_state()
    st.set_page_config(
        page_title=STREAMLIT_APP_TITLE,
        layout="wide",
        initial_sidebar_state="expanded",
    )
    if "streamlit_theme_name" not in st.session_state:
        st.session_state["streamlit_theme_name"] = _normalize_streamlit_theme_name(
            persisted_boot_state.get("theme_name", DEFAULT_STREAMLIT_THEME)
        )
    active_theme_name = _normalize_streamlit_theme_name(
        st.session_state.get("streamlit_theme_name")
    )
    st.session_state["streamlit_theme_name"] = active_theme_name
    st.markdown(_build_streamlit_theme_css(active_theme_name), unsafe_allow_html=True)

db_path = DB_PATH_DEFAULT
docs_dir = DOCS_DIR_DEFAULT
consult_api = False

if REAL_RUNTIME:
    if "streamlit_source_state" not in st.session_state:
        st.session_state["streamlit_source_state"] = {
            "db_path": DB_PATH_DEFAULT,
            "docs_dir": DOCS_DIR_DEFAULT,
        }
    source_state = st.session_state["streamlit_source_state"]
    db_path_candidate = str(source_state.get("db_path", DB_PATH_DEFAULT))
    docs_dir_candidate = str(source_state.get("docs_dir", DOCS_DIR_DEFAULT))
    try:
        db_path = str(
            ensure_path_is_allowed(
                db_path_candidate,
                purpose="Arquivo do banco",
                expect_directory=False,
            )
        )
        docs_dir = str(
            ensure_path_is_allowed(
                docs_dir_candidate,
                purpose="Pasta com planilhas",
                expect_directory=True,
            )
        )
    except PathSafetyError:
        db_path = DB_PATH_DEFAULT
        docs_dir = DOCS_DIR_DEFAULT
        st.session_state["streamlit_source_state"] = {
            "db_path": db_path,
            "docs_dir": docs_dir,
        }

    header_left, header_right = st.columns([5, 1])
    with header_left:
        st.markdown(f"<h1>{STREAMLIT_APP_TITLE}</h1>", unsafe_allow_html=True)
        st.markdown(
            "<p class='section-label'>Painel streamlit com filtros, tabela paginada e exportacao</p>",
            unsafe_allow_html=True,
        )
    with header_right:
        theme_options = list(STREAMLIT_THEME_PALETTES.keys())
        selected_theme_header = st.selectbox(
            "Tema",
            theme_options,
            index=theme_options.index(active_theme_name),
            key="streamlit_theme_header_selector",
        )
        if selected_theme_header != active_theme_name:
            st.session_state["streamlit_theme_name"] = selected_theme_header
            _persist_streamlit_state(
                theme_name=selected_theme_header,
                width_profile="Padrao (1600)",
                width_profile_by_bucket={},
                streamlit_render_stats=_normalize_render_stats(
                    st.session_state.get("streamlit_render_stats", {})
                ),
            )
            rerun_fn = getattr(st, "rerun", None)
            if callable(rerun_fn):
                rerun_fn()
        st.caption(f"Atualizado as {datetime.now().strftime('%H:%M:%S')}")

    with st.sidebar:
        st.subheader("Painel rapido")
        db_file = Path(db_path)
        db_exists = db_file.exists()
        db_size_mb = (db_file.stat().st_size / (1024 * 1024)) if db_exists else 0.0
        sheet_files_count = _list_spreadsheet_files_count(docs_dir)
        st.caption(f"DB: {'ok' if db_exists else 'ausente'} | {db_size_mb:.1f} MB")
        st.caption(f"Arquivos de entrada: {sheet_files_count}")

raw_df = load_dataframe(db_path) if REAL_RUNTIME else pd.DataFrame()
sidebar_filtered_metric_slot: Any = None
base_derivada_context_map: Optional[pd.DataFrame] = None
if REAL_RUNTIME and raw_df.empty:
    st.info(
        "Banco nao encontrado ou sem dados. Use a aba 'Cache e API' em 'Fonte de dados avancada'."
    )
    st.stop()

if REAL_RUNTIME and not raw_df.empty:
    if "numero_ssa" in raw_df.columns and "derivada_de" in raw_df.columns:
        try:
            base_derivada_context_map = _build_derivada_context_map(raw_df)
        except Exception as exc:
            logger.debug(
                "Falha ao precomputar contexto de derivadas no Streamlit: %s", exc
            )
            base_derivada_context_map = None
    available_columns_runtime = _columns_with_data(raw_df, list(raw_df.columns))
    if not available_columns_runtime:
        available_columns_runtime = list(raw_df.columns)
    with st.sidebar:
        st.divider()
        st.subheader("Resumo rapido")
        weekly_kpis = _compute_sidebar_weekly_kpis_cached(
            raw_df,
            datetime.now().strftime("%Y-%m-%d"),
        )
        st.metric("Registros no banco", len(raw_df))
        sidebar_filtered_metric_slot = st.empty()
        exec_col, emit_col = st.columns(2)
        exec_col.metric("Exec semana atual", weekly_kpis["executadas_semana_atual"])
        exec_col.metric(
            "Exec semana anterior", weekly_kpis["executadas_semana_anterior"]
        )
        emit_col.metric("Emit semana atual", weekly_kpis["emitidas_semana_atual"])
        emit_col.metric("Emit semana anterior", weekly_kpis["emitidas_semana_anterior"])

search_terms = ""
situacao_sel: list[Any] = []
executor_sel: list[Any] = []
emissor_sel: list[Any] = []
executor_resp_sel: list[Any] = []
estado_sel: list[Any] = []
ano_emissao_sel: list[Any] = []
ano_execucao_sel: list[Any] = []
ano_semana_emissao_sel: list[Any] = []
ano_semana_execucao_sel: list[Any] = []
num_reprogramacoes_sel: list[Any] = []
data_emissao_inicio = ""
data_emissao_fim = ""
data_execucao_inicio = ""
data_execucao_fim = ""
tem_derivada_mode = "todos"
tem_derivadas_mode = "todos"
selected_columns = list(raw_df.columns)
limit_rows = 500
page_size = 250
page_number = 1
table_height = 600
auto_width = True
recent_df: pd.DataFrame | None = None
situacoes: list[Any] = []
executores: list[Any] = []
emissores: list[Any] = []

if REAL_RUNTIME and not raw_df.empty:
    persisted_streamlit_state = _load_persisted_streamlit_state()
    tab_filters, tab_table, tab_export, tab_ops = st.tabs(MAIN_TAB_LABELS)

    with tab_filters:
        st.subheader("Filtros")
        situacoes, executores, emissores = _build_filter_options(raw_df)
        advanced_options = _build_advanced_filter_options(raw_df)
        default_executor: list[Any] = []
        default_emissor: list[Any] = []
        available_columns = _columns_with_data(raw_df, list(raw_df.columns))
        if not available_columns:
            available_columns = list(raw_df.columns)
        column_display_names = {
            col: DISPLAY_MAPPINGS.get(col, col) for col in available_columns
        }
        default_columns = _default_visible_columns(list(available_columns))
        column_presets = _build_column_presets(list(available_columns))

        state_key = "streamlit_filters_state"
        if state_key not in st.session_state:
            st.session_state[state_key] = {
                "search_terms": "",
                "consult_api": False,
                "situacao_sel": [],
                "executor_sel": default_executor,
                "emissor_sel": default_emissor,
                "executor_resp_sel": [],
                "executor_resp_manual": "",
                "executor_resp_exclude": [],
                "estado_sel": [],
                "estado_manual": "",
                "estado_exclude": [],
                "ano_emissao_sel": [],
                "ano_emissao_manual": "",
                "ano_emissao_exclude": [],
                "ano_execucao_sel": [],
                "ano_execucao_manual": "",
                "ano_execucao_exclude": [],
                "ano_semana_emissao_sel": [],
                "ano_semana_emissao_manual": "",
                "ano_semana_emissao_exclude": [],
                "ano_semana_execucao_sel": [],
                "ano_semana_execucao_manual": "",
                "ano_semana_execucao_exclude": [],
                "num_reprogramacoes_sel": [],
                "num_reprogramacoes_manual": "",
                "num_reprogramacoes_exclude": [],
                "data_emissao_inicio": "",
                "data_emissao_fim": "",
                "data_execucao_inicio": "",
                "data_execucao_fim": "",
                "tem_derivada_mode": "todos",
                "tem_derivadas_mode": "todos",
                "use_calendar_mode": False,
                "limit_rows": limit_rows,
                "selected_display": [
                    column_display_names[col] for col in default_columns
                ],
            }
        filter_state = st.session_state[state_key]
        filter_state.setdefault("executor_resp_exclude", [])
        filter_state.setdefault("estado_exclude", [])
        filter_state.setdefault("ano_emissao_exclude", [])
        filter_state.setdefault("ano_execucao_exclude", [])
        filter_state.setdefault("ano_semana_emissao_exclude", [])
        filter_state.setdefault("ano_semana_execucao_exclude", [])
        filter_state.setdefault("num_reprogramacoes_exclude", [])
        filter_state["situacao_sel"] = [
            value
            for value in filter_state.get("situacao_sel", [])
            if value in situacoes
        ]
        filter_state["executor_sel"] = [
            value
            for value in filter_state.get("executor_sel", [])
            if value in executores
        ]
        filter_state["emissor_sel"] = [
            value for value in filter_state.get("emissor_sel", []) if value in emissores
        ]
        filter_state["executor_resp_sel"] = [
            value
            for value in filter_state.get("executor_resp_sel", [])
            if value in advanced_options.get("executor_resp", [])
        ]
        filter_state["estado_sel"] = [
            value
            for value in filter_state.get("estado_sel", [])
            if value in advanced_options.get("estado", [])
        ]
        filter_state["ano_emissao_sel"] = [
            value
            for value in filter_state.get("ano_emissao_sel", [])
            if value in advanced_options.get("ano_emissao", [])
        ]
        filter_state["ano_execucao_sel"] = [
            value
            for value in filter_state.get("ano_execucao_sel", [])
            if value in advanced_options.get("ano_execucao", [])
        ]
        filter_state["ano_semana_emissao_sel"] = [
            value
            for value in filter_state.get("ano_semana_emissao_sel", [])
            if value in advanced_options.get("ano_semana_emissao", [])
        ]
        filter_state["ano_semana_execucao_sel"] = [
            value
            for value in filter_state.get("ano_semana_execucao_sel", [])
            if value in advanced_options.get("ano_semana_execucao", [])
        ]
        filter_state["num_reprogramacoes_sel"] = [
            value
            for value in filter_state.get("num_reprogramacoes_sel", [])
            if value in advanced_options.get("num_reprogramacoes", [])
        ]
        filter_state["data_emissao_inicio"] = str(
            filter_state.get("data_emissao_inicio", "") or ""
        )
        filter_state["data_emissao_fim"] = str(
            filter_state.get("data_emissao_fim", "") or ""
        )
        filter_state["data_execucao_inicio"] = str(
            filter_state.get("data_execucao_inicio", "") or ""
        )
        filter_state["data_execucao_fim"] = str(
            filter_state.get("data_execucao_fim", "") or ""
        )
        if str(filter_state.get("tem_derivada_mode", "todos")).lower() not in {
            "todos",
            "sim",
            "nao",
        }:
            filter_state["tem_derivada_mode"] = "todos"
        if str(filter_state.get("tem_derivadas_mode", "todos")).lower() not in {
            "todos",
            "sim",
            "nao",
        }:
            filter_state["tem_derivadas_mode"] = "todos"
        filter_state["use_calendar_mode"] = bool(
            filter_state.get("use_calendar_mode", False)
        )
        valid_display_values = set(column_display_names.values())
        selected_display_state = [
            value
            for value in filter_state.get("selected_display", [])
            if value in valid_display_values
        ]
        filter_state["selected_display"] = selected_display_state or [
            column_display_names[col] for col in default_columns
        ]

        table_state_key = "streamlit_table_state"
        if table_state_key not in st.session_state:
            st.session_state[table_state_key] = {
                "sort_column": "(Sem ordenacao)",
                "sort_desc": False,
                "page_size": 100,
                "table_height": 600,
                "auto_width": True,
                "page_number": 1,
                "width_profile": str(
                    persisted_streamlit_state.get("width_profile", "Padrao (1600)")
                ),
                "width_profile_by_bucket": _normalize_width_profile_memory(
                    persisted_streamlit_state.get("width_profile_by_bucket", {})
                ),
                "table_mode": "Tabela + grafico",
                "compact_mode": False,
            }
        if "streamlit_render_stats" not in st.session_state:
            st.session_state["streamlit_render_stats"] = _normalize_render_stats(
                persisted_streamlit_state.get("streamlit_render_stats", {})
            )
        table_state = st.session_state[table_state_key]
        table_state["page_size"] = int(table_state.get("page_size", 100))
        table_state["table_height"] = int(table_state.get("table_height", 600))
        table_state["auto_width"] = bool(table_state.get("auto_width", True))
        table_state["sort_desc"] = bool(table_state.get("sort_desc", False))
        table_state["page_number"] = int(table_state.get("page_number", 1))
        table_state["width_profile"] = str(
            table_state.get("width_profile", "Padrao (1600)")
        )
        table_state["width_profile_by_bucket"] = _normalize_width_profile_memory(
            table_state.get("width_profile_by_bucket", {})
        )
        table_state["table_mode"] = str(
            table_state.get("table_mode", "Tabela + grafico")
        )
        table_state["compact_mode"] = bool(table_state.get("compact_mode", False))

        with st.form("filters_form", clear_on_submit=False):
            st.caption("Busca e origem")
            search_row = st.columns([4.5, 1.2])
            search_input = search_row[0].text_input(
                "Busca (mesma sintaxe da CLI)",
                value=filter_state.get("search_terms", ""),
                placeholder="ex.: svp, !ste, mel4",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            apply_search_now = search_row[1].form_submit_button("Filtrar agora")
            consult_api_input = st.checkbox(
                "Consulta API manual",
                value=bool(filter_state.get("consult_api", False)),
            )
            st.caption("Filtros principais")
            row_filters = st.columns([0.9, 0.9, 1.2, 0.7])
            executor_input_multi = row_filters[0].multiselect(
                "Setor exec",
                options=executores,
                default=filter_state.get("executor_sel", default_executor),
            )
            emissor_input_multi = row_filters[1].multiselect(
                "Setor emissor",
                options=emissores,
                default=filter_state.get("emissor_sel", default_emissor),
            )
            situacao_input = row_filters[2].multiselect(
                "Estado",
                options=situacoes,
                default=filter_state.get("situacao_sel", []),
            )
            limit_input = int(
                row_filters[3].number_input(
                    "Limite",
                    min_value=50,
                    max_value=20000,
                    value=int(filter_state.get("limit_rows", limit_rows)),
                    step=50,
                )
            )
            st.caption("Filtros avancados")
            advanced_row_1 = st.columns(4)
            executor_resp_input = advanced_row_1[0].multiselect(
                "Executor",
                options=advanced_options.get("executor_resp", []),
                default=filter_state.get("executor_resp_sel", []),
            )
            estado_input = advanced_row_1[1].multiselect(
                "Estado detalhado",
                options=advanced_options.get("estado", []),
                default=filter_state.get("estado_sel", []),
            )
            ano_emissao_input = advanced_row_1[2].multiselect(
                "Ano emissao",
                options=advanced_options.get("ano_emissao", []),
                default=filter_state.get("ano_emissao_sel", []),
            )
            ano_execucao_input = advanced_row_1[3].multiselect(
                "Ano execucao",
                options=advanced_options.get("ano_execucao", []),
                default=filter_state.get("ano_execucao_sel", []),
            )
            # Streamlit sidequest policy: keep filter state/UI centralized in this single file.
            advanced_row_1_manual = st.columns(4)
            executor_resp_manual_input = advanced_row_1_manual[0].text_input(
                "Executor (! exclui)",
                value=str(filter_state.get("executor_resp_manual", "")),
                placeholder="ex.: op1,!op2",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            estado_manual_input = advanced_row_1_manual[1].text_input(
                "Estado detalhado (! exclui)",
                value=str(filter_state.get("estado_manual", "")),
                placeholder="ex.: ste,!ate",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            ano_emissao_manual_input = advanced_row_1_manual[2].text_input(
                "Ano emissao (! exclui)",
                value=str(filter_state.get("ano_emissao_manual", "")),
                placeholder="ex.: 2024,!2023",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            ano_execucao_manual_input = advanced_row_1_manual[3].text_input(
                "Ano execucao (! exclui)",
                value=str(filter_state.get("ano_execucao_manual", "")),
                placeholder="ex.: 2025,!2022",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            advanced_row_2 = st.columns(4)
            ano_semana_emissao_input = advanced_row_2[0].multiselect(
                "AnoSemana emissao",
                options=advanced_options.get("ano_semana_emissao", []),
                default=filter_state.get("ano_semana_emissao_sel", []),
            )
            ano_semana_execucao_input = advanced_row_2[1].multiselect(
                "AnoSemana execucao",
                options=advanced_options.get("ano_semana_execucao", []),
                default=filter_state.get("ano_semana_execucao_sel", []),
            )
            num_reprogramacoes_input = advanced_row_2[2].multiselect(
                "Num reprogramacoes",
                options=advanced_options.get("num_reprogramacoes", []),
                default=filter_state.get("num_reprogramacoes_sel", []),
            )
            deriv_mode_options = ["todos", "sim", "nao"]
            tem_derivada_input = advanced_row_2[3].selectbox(
                "Tem derivada",
                options=deriv_mode_options,
                index=deriv_mode_options.index(
                    str(filter_state.get("tem_derivada_mode", "todos"))
                ),
            )
            advanced_row_2_manual = st.columns(4)
            ano_semana_emissao_manual_input = advanced_row_2_manual[0].text_input(
                "AnoSemana emissao (! exclui)",
                value=str(filter_state.get("ano_semana_emissao_manual", "")),
                placeholder="ex.: 202401,!202352",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            ano_semana_execucao_manual_input = advanced_row_2_manual[1].text_input(
                "AnoSemana execucao (! exclui)",
                value=str(filter_state.get("ano_semana_execucao_manual", "")),
                placeholder="ex.: 202410,!202409",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            num_reprogramacoes_manual_input = advanced_row_2_manual[2].text_input(
                "Num reprogramacoes (! exclui)",
                value=str(filter_state.get("num_reprogramacoes_manual", "")),
                placeholder="ex.: 0,!2",
                help="Use virgula para multiplos valores e ! para diferente de.",
            )
            use_calendar_mode = st.checkbox(
                "Usar calendario nas datas",
                value=bool(filter_state.get("use_calendar_mode", False)),
                help="Desative para usar sintaxe manual com ! e virgula.",
            )
            advanced_row_3 = st.columns(4)
            data_emissao_inicio_manual = advanced_row_3[0].text_input(
                "Data emissao inicio",
                value=str(filter_state.get("data_emissao_inicio", "")),
                placeholder="YYYY-MM-DD ou !YYYY-MM-DD",
                help="Campo manual aceita sintaxe com ! e virgula.",
                disabled=use_calendar_mode,
            )
            data_emissao_inicio_picker = advanced_row_3[0].date_input(
                "Calendario emissao inicio",
                # Keep None when raw input is empty to avoid injecting implicit date filters.
                value=_parse_date_input(filter_state.get("data_emissao_inicio", "")),
                format="YYYY-MM-DD",
                key="cal_data_emissao_inicio",
                disabled=not use_calendar_mode,
            )
            data_emissao_fim_manual = advanced_row_3[1].text_input(
                "Data emissao fim",
                value=str(filter_state.get("data_emissao_fim", "")),
                placeholder="YYYY-MM-DD ou !YYYY-MM-DD",
                help="Campo manual aceita sintaxe com ! e virgula.",
                disabled=use_calendar_mode,
            )
            data_emissao_fim_picker = advanced_row_3[1].date_input(
                "Calendario emissao fim",
                # Keep None when raw input is empty to avoid injecting implicit date filters.
                value=_parse_date_input(filter_state.get("data_emissao_fim", "")),
                format="YYYY-MM-DD",
                key="cal_data_emissao_fim",
                disabled=not use_calendar_mode,
            )
            data_execucao_inicio_manual = advanced_row_3[2].text_input(
                "Data execucao inicio",
                value=str(filter_state.get("data_execucao_inicio", "")),
                placeholder="YYYY-MM-DD ou !YYYY-MM-DD",
                help="Campo manual aceita sintaxe com ! e virgula.",
                disabled=use_calendar_mode,
            )
            data_execucao_inicio_picker = advanced_row_3[2].date_input(
                "Calendario execucao inicio",
                # Keep None when raw input is empty to avoid injecting implicit date filters.
                value=_parse_date_input(filter_state.get("data_execucao_inicio", "")),
                format="YYYY-MM-DD",
                key="cal_data_execucao_inicio",
                disabled=not use_calendar_mode,
            )
            data_execucao_fim_manual = advanced_row_3[3].text_input(
                "Data execucao fim",
                value=str(filter_state.get("data_execucao_fim", "")),
                placeholder="YYYY-MM-DD ou !YYYY-MM-DD",
                help="Campo manual aceita sintaxe com ! e virgula.",
                disabled=use_calendar_mode,
            )
            data_execucao_fim_picker = advanced_row_3[3].date_input(
                "Calendario execucao fim",
                # Keep None when raw input is empty to avoid injecting implicit date filters.
                value=_parse_date_input(filter_state.get("data_execucao_fim", "")),
                format="YYYY-MM-DD",
                key="cal_data_execucao_fim",
                disabled=not use_calendar_mode,
            )
            tem_derivadas_input = st.selectbox(
                "Tem derivadas (estrutura)",
                options=deriv_mode_options,
                index=deriv_mode_options.index(
                    str(filter_state.get("tem_derivadas_mode", "todos"))
                ),
            )
            st.caption("Colunas visiveis")
            st.caption(
                "Essas colunas sao refletidas na aba Tabela e podem ser ajustadas la tambem."
            )
            display_options = [column_display_names[col] for col in available_columns]
            selected_display_base = filter_state.get(
                "selected_display",
                [column_display_names[col] for col in default_columns],
            )
            mark_all_columns = st.checkbox(
                "Marcar tudo",
                value=len(selected_display_base) >= len(display_options),
            )
            selected_display_input = st.multiselect(
                "Colunas exibidas",
                options=display_options,
                default=display_options if mark_all_columns else selected_display_base,
            )
            if mark_all_columns:
                selected_display_input = display_options
            preset_cols = st.columns(3)
            preset_core = preset_cols[0].form_submit_button("Operacao diaria")
            preset_all = preset_cols[1].form_submit_button("Analise completa")
            preset_min = preset_cols[2].form_submit_button("Minimo")
            with st.expander("Ajuda rapida", expanded=False):
                st.markdown(
                    "* Sintaxe basica: `svp, !ste, mel4`\n"
                    "* Use `OU` ou `OR` para alternativas (`svp OU mel4`)\n"
                    "* Prefixos uteis: `^` inicio, `$` final, `=` igual, `~` regex\n"
                    "* `!` inverte termo (`!^adm`, `!mel4`)\n"
                    "* Virgulas equivalem a E/AND; espacos tambem separam termos"
                )
            form_cols = st.columns(2)
            apply_filters = form_cols[0].form_submit_button("Aplicar filtros")
            reset_filters = form_cols[1].form_submit_button("Resetar filtros")

        if reset_filters:
            for cal_key in (
                "cal_data_emissao_inicio",
                "cal_data_emissao_fim",
                "cal_data_execucao_inicio",
                "cal_data_execucao_fim",
            ):
                st.session_state.pop(cal_key, None)
            st.session_state[state_key] = {
                "search_terms": "",
                "consult_api": False,
                "situacao_sel": [],
                "executor_sel": [],
                "emissor_sel": [],
                "executor_resp_sel": [],
                "executor_resp_manual": "",
                "executor_resp_exclude": [],
                "estado_sel": [],
                "estado_manual": "",
                "estado_exclude": [],
                "ano_emissao_sel": [],
                "ano_emissao_manual": "",
                "ano_emissao_exclude": [],
                "ano_execucao_sel": [],
                "ano_execucao_manual": "",
                "ano_execucao_exclude": [],
                "ano_semana_emissao_sel": [],
                "ano_semana_emissao_manual": "",
                "ano_semana_emissao_exclude": [],
                "ano_semana_execucao_sel": [],
                "ano_semana_execucao_manual": "",
                "ano_semana_execucao_exclude": [],
                "num_reprogramacoes_sel": [],
                "num_reprogramacoes_manual": "",
                "num_reprogramacoes_exclude": [],
                "data_emissao_inicio": "",
                "data_emissao_fim": "",
                "data_execucao_inicio": "",
                "data_execucao_fim": "",
                "tem_derivada_mode": "todos",
                "tem_derivadas_mode": "todos",
                "use_calendar_mode": False,
                "limit_rows": 500,
                "selected_display": [
                    column_display_names[col] for col in default_columns
                ],
            }
            st.session_state[table_state_key] = {
                "sort_column": "(Sem ordenacao)",
                "sort_desc": False,
                "page_size": 100,
                "table_height": 600,
                "auto_width": True,
                "page_number": 1,
                "width_profile": "Padrao (1600)",
                "width_profile_by_bucket": {},
                "table_mode": "Tabela + grafico",
                "compact_mode": False,
            }
            _persist_streamlit_state(
                width_profile="Padrao (1600)",
                width_profile_by_bucket={},
                streamlit_render_stats=_normalize_render_stats(
                    st.session_state.get("streamlit_render_stats", {})
                ),
            )
            rerun_fn = getattr(st, "rerun", None)
            if callable(rerun_fn):
                rerun_fn()
            else:
                legacy_rerun_fn = getattr(st, "experimental_rerun", None)
                if callable(legacy_rerun_fn):
                    legacy_rerun_fn()

        if preset_core:
            filter_state["selected_display"] = [
                column_display_names[col] for col in column_presets["core"]
            ]
        elif preset_all:
            filter_state["selected_display"] = [
                column_display_names[col] for col in column_presets["all"]
            ]
        elif preset_min:
            min_cols = [
                col
                for col in ("numero_ssa", "situacao", "descricao_ssa")
                if col in available_columns
            ]
            filter_state["selected_display"] = [
                column_display_names[col] for col in min_cols
            ]

        if apply_filters or apply_search_now:
            st.session_state[state_key] = {
                "search_terms": search_input,
                "consult_api": consult_api_input,
                "situacao_sel": list(situacao_input),
                "executor_sel": list(executor_input_multi),
                "emissor_sel": list(emissor_input_multi),
                "executor_resp_sel": list(executor_resp_input),
                "executor_resp_manual": str(executor_resp_manual_input or "").strip(),
                "executor_resp_exclude": [],
                "estado_sel": list(estado_input),
                "estado_manual": str(estado_manual_input or "").strip(),
                "estado_exclude": [],
                "ano_emissao_sel": list(ano_emissao_input),
                "ano_emissao_manual": str(ano_emissao_manual_input or "").strip(),
                "ano_emissao_exclude": [],
                "ano_execucao_sel": list(ano_execucao_input),
                "ano_execucao_manual": str(ano_execucao_manual_input or "").strip(),
                "ano_execucao_exclude": [],
                "ano_semana_emissao_sel": list(ano_semana_emissao_input),
                "ano_semana_emissao_manual": str(
                    ano_semana_emissao_manual_input or ""
                ).strip(),
                "ano_semana_emissao_exclude": [],
                "ano_semana_execucao_sel": list(ano_semana_execucao_input),
                "ano_semana_execucao_manual": str(
                    ano_semana_execucao_manual_input or ""
                ).strip(),
                "ano_semana_execucao_exclude": [],
                "num_reprogramacoes_sel": list(num_reprogramacoes_input),
                "num_reprogramacoes_manual": str(
                    num_reprogramacoes_manual_input or ""
                ).strip(),
                "num_reprogramacoes_exclude": [],
                "data_emissao_inicio": (
                    _format_date_input(data_emissao_inicio_picker)
                    if use_calendar_mode
                    else str(data_emissao_inicio_manual or "").strip()
                ),
                "data_emissao_fim": (
                    _format_date_input(data_emissao_fim_picker)
                    if use_calendar_mode
                    else str(data_emissao_fim_manual or "").strip()
                ),
                "data_execucao_inicio": (
                    _format_date_input(data_execucao_inicio_picker)
                    if use_calendar_mode
                    else str(data_execucao_inicio_manual or "").strip()
                ),
                "data_execucao_fim": (
                    _format_date_input(data_execucao_fim_picker)
                    if use_calendar_mode
                    else str(data_execucao_fim_manual or "").strip()
                ),
                "tem_derivada_mode": str(tem_derivada_input),
                "tem_derivadas_mode": str(tem_derivadas_input),
                "use_calendar_mode": bool(use_calendar_mode),
                "limit_rows": limit_input,
                "selected_display": selected_display_input,
            }

        filter_state = st.session_state[state_key]
        search_terms = str(filter_state.get("search_terms", ""))
        consult_api = bool(filter_state.get("consult_api", False))
        situacao_sel = list(filter_state.get("situacao_sel", []))
        executor_sel = list(filter_state.get("executor_sel", []))
        emissor_sel = list(filter_state.get("emissor_sel", []))
        executor_resp_sel = list(filter_state.get("executor_resp_sel", []))
        estado_sel = list(filter_state.get("estado_sel", []))
        ano_emissao_sel = list(filter_state.get("ano_emissao_sel", []))
        ano_execucao_sel = list(filter_state.get("ano_execucao_sel", []))
        ano_semana_emissao_sel = list(filter_state.get("ano_semana_emissao_sel", []))
        ano_semana_execucao_sel = list(filter_state.get("ano_semana_execucao_sel", []))
        num_reprogramacoes_sel = list(filter_state.get("num_reprogramacoes_sel", []))
        data_emissao_inicio = str(filter_state.get("data_emissao_inicio", "") or "")
        data_emissao_fim = str(filter_state.get("data_emissao_fim", "") or "")
        data_execucao_inicio = str(filter_state.get("data_execucao_inicio", "") or "")
        data_execucao_fim = str(filter_state.get("data_execucao_fim", "") or "")
        tem_derivada_mode = str(filter_state.get("tem_derivada_mode", "todos"))
        tem_derivadas_mode = str(filter_state.get("tem_derivadas_mode", "todos"))
        limit_rows = int(filter_state.get("limit_rows", 500))
        selected_display = list(
            filter_state.get(
                "selected_display",
                [column_display_names[col] for col in default_columns],
            )
        )
        display_to_internal = {v: k for k, v in column_display_names.items()}
        selected_columns = [
            display_to_internal.get(name, name) for name in selected_display
        ]

    situacao_filter = _normalize_filter_selection(situacao_sel, situacoes)
    executor_filter = _normalize_filter_selection(executor_sel, executores)
    emissor_filter = _normalize_filter_selection(emissor_sel, emissores)
    executor_resp_filter, executor_resp_exclude = _resolve_include_exclude_values(
        executor_resp_sel,
        advanced_options.get("executor_resp", []),
        filter_state.get("executor_resp_manual", ""),
    )
    estado_filter, estado_exclude = _resolve_include_exclude_values(
        estado_sel,
        advanced_options.get("estado", []),
        filter_state.get("estado_manual", ""),
    )
    ano_emissao_filter, ano_emissao_exclude = _resolve_include_exclude_values(
        ano_emissao_sel,
        advanced_options.get("ano_emissao", []),
        filter_state.get("ano_emissao_manual", ""),
    )
    ano_execucao_filter, ano_execucao_exclude = _resolve_include_exclude_values(
        ano_execucao_sel,
        advanced_options.get("ano_execucao", []),
        filter_state.get("ano_execucao_manual", ""),
    )
    ano_semana_emissao_filter, ano_semana_emissao_exclude = (
        _resolve_include_exclude_values(
            ano_semana_emissao_sel,
            advanced_options.get("ano_semana_emissao", []),
            filter_state.get("ano_semana_emissao_manual", ""),
        )
    )
    ano_semana_execucao_filter, ano_semana_execucao_exclude = (
        _resolve_include_exclude_values(
            ano_semana_execucao_sel,
            advanced_options.get("ano_semana_execucao", []),
            filter_state.get("ano_semana_execucao_manual", ""),
        )
    )
    num_reprogramacoes_filter, num_reprogramacoes_exclude = (
        _resolve_include_exclude_values(
            num_reprogramacoes_sel,
            advanced_options.get("num_reprogramacoes", []),
            filter_state.get("num_reprogramacoes_manual", ""),
        )
    )
    filter_state["executor_resp_exclude"] = list(executor_resp_exclude)
    filter_state["estado_exclude"] = list(estado_exclude)
    filter_state["ano_emissao_exclude"] = list(ano_emissao_exclude)
    filter_state["ano_execucao_exclude"] = list(ano_execucao_exclude)
    filter_state["ano_semana_emissao_exclude"] = list(ano_semana_emissao_exclude)
    filter_state["ano_semana_execucao_exclude"] = list(ano_semana_execucao_exclude)
    filter_state["num_reprogramacoes_exclude"] = list(num_reprogramacoes_exclude)
    advanced_filters = {
        "executor_resp": executor_resp_filter,
        "executor_resp_exclude": executor_resp_exclude,
        "estado": estado_filter,
        "estado_exclude": estado_exclude,
        "ano_emissao": ano_emissao_filter,
        "ano_emissao_exclude": ano_emissao_exclude,
        "ano_execucao": ano_execucao_filter,
        "ano_execucao_exclude": ano_execucao_exclude,
        "ano_semana_emissao": ano_semana_emissao_filter,
        "ano_semana_emissao_exclude": ano_semana_emissao_exclude,
        "ano_semana_execucao": ano_semana_execucao_filter,
        "ano_semana_execucao_exclude": ano_semana_execucao_exclude,
        "num_reprogramacoes": num_reprogramacoes_filter,
        "num_reprogramacoes_exclude": num_reprogramacoes_exclude,
        "data_emissao_inicio": data_emissao_inicio,
        "data_emissao_fim": data_emissao_fim,
        "data_execucao_inicio": data_execucao_inicio,
        "data_execucao_fim": data_execucao_fim,
        "tem_derivada": str(tem_derivada_mode).strip().lower(),
        "tem_derivadas": str(tem_derivadas_mode).strip().lower(),
    }

    filtered_df = apply_all_filters_cached(
        raw_df,
        search_terms,
        situacao_filter,
        executor_filter,
        emissor_filter,
        advanced_filters=advanced_filters,
        base_derivada_context_map=base_derivada_context_map,
    )
    if limit_rows and len(filtered_df) > limit_rows:
        filtered_df = filtered_df.head(limit_rows).reset_index(drop=True)
    if sidebar_filtered_metric_slot is not None:
        sidebar_filtered_metric_slot.metric("SSAs no filtro atual", len(filtered_df))

    view_df = filtered_df[selected_columns] if selected_columns else filtered_df
    rename_map = {col: DISPLAY_MAPPINGS.get(col, col) for col in view_df.columns}
    rename_map.update(
        {
            "_tem_derivada_ctx": "Tem derivada",
            "_tem_derivadas_ctx": "Tem derivadas",
            "_qtd_derivadas_ctx": "Qtd derivadas",
        }
    )

    active_summary = _build_active_summary(
        search_terms=search_terms,
        situacao_filter=situacao_filter,
        executor_filter=executor_filter,
        emissor_filter=emissor_filter,
        advanced_filters=advanced_filters,
        consult_api=consult_api,
    )

    with tab_table:
        with st.expander("Colunas exibidas (atalho rapido)", expanded=False):
            quick_display_options = [
                column_display_names[col] for col in available_columns
            ]
            quick_selected_default = filter_state.get(
                "selected_display",
                [column_display_names[col] for col in default_columns],
            )
            quick_mark_all = st.checkbox(
                "Marcar tudo (tabela)",
                value=len(quick_selected_default) >= len(quick_display_options),
                key="table_quick_mark_all",
            )
            quick_selected_display = st.multiselect(
                "Colunas da tabela",
                options=quick_display_options,
                default=quick_display_options
                if quick_mark_all
                else quick_selected_default,
                key="table_quick_selected_display",
            )
            if quick_mark_all:
                quick_selected_display = quick_display_options
            quick_cols = st.columns(3)
            quick_core = quick_cols[0].button("Operacao diaria", key="table_quick_core")
            quick_all = quick_cols[1].button("Analise completa", key="table_quick_all")
            quick_apply = quick_cols[2].button("Aplicar", key="table_quick_apply")

            if quick_core:
                st.session_state[state_key]["selected_display"] = [
                    column_display_names[col] for col in column_presets["core"]
                ]
                rerun_fn = getattr(st, "rerun", None)
                if callable(rerun_fn):
                    rerun_fn()
            elif quick_all:
                st.session_state[state_key]["selected_display"] = [
                    column_display_names[col] for col in column_presets["all"]
                ]
                rerun_fn = getattr(st, "rerun", None)
                if callable(rerun_fn):
                    rerun_fn()
            elif quick_apply:
                st.session_state[state_key]["selected_display"] = quick_selected_display
                rerun_fn = getattr(st, "rerun", None)
                if callable(rerun_fn):
                    rerun_fn()

        total_ssas = len(filtered_df)
        original_count = len(raw_df)
        status_cols = st.columns(3)
        status_cols[0].metric("Total filtrado", total_ssas)
        status_cols[1].metric("Total banco", original_count)
        status_cols[2].metric("Colunas exibidas", len(view_df.columns))

        if "situacao" in filtered_df.columns and total_ssas:
            status_counts = filtered_df["situacao"].value_counts()
            executadas = int(status_counts.get("EXECUTADA", 0))
            exec_rate = (executadas / total_ssas * 100) if total_ssas else 0
            st.caption(f"Execucao concluida: {exec_rate:.1f}%")
        else:
            st.caption("Execucao concluida: -")

        info_cols = st.columns(3)
        info_cols[0].metric(
            "Situacoes distintas",
            int(filtered_df["situacao"].nunique())
            if "situacao" in filtered_df.columns
            else 0,
        )
        info_cols[1].metric(
            "Executores distintos",
            int(filtered_df["setor_executor"].nunique())
            if "setor_executor" in filtered_df.columns
            else 0,
        )
        info_cols[2].metric(
            "Emissores distintos",
            int(filtered_df["setor_emissor"].nunique())
            if "setor_emissor" in filtered_df.columns
            else 0,
        )

        primary_controls = st.columns([2.0, 0.9, 1.3, 1.0])
        sort_options = ["(Sem ordenacao)"] + list(view_df.columns)
        sort_column = str(
            primary_controls[0].selectbox(
                "Ordenar por",
                sort_options,
                index=sort_options.index(
                    table_state.get("sort_column", "(Sem ordenacao)")
                )
                if table_state.get("sort_column", "(Sem ordenacao)") in sort_options
                else 0,
            )
        )
        sort_desc = bool(
            primary_controls[1].checkbox(
                "Desc", value=table_state.get("sort_desc", False)
            )
        )
        table_mode_options = ["Tabela", "Tabela + grafico"]
        table_mode = str(
            primary_controls[2].radio(
                "Visualizacao",
                options=table_mode_options,
                index=table_mode_options.index(
                    table_state.get("table_mode", "Tabela + grafico")
                )
                if table_state.get("table_mode", "Tabela + grafico")
                in table_mode_options
                else 1,
                horizontal=True,
            )
        )
        compact_mode = bool(
            primary_controls[3].checkbox(
                "Compacto",
                value=table_state.get("compact_mode", False),
            )
        )

        secondary_controls = st.columns([1.3, 1.2, 1.1, 1.5])
        page_size = int(
            secondary_controls[0].selectbox(
                "Linhas por pagina",
                [25, 50, 100, 250, 500],
                index=[25, 50, 100, 250, 500].index(table_state.get("page_size", 100))
                if table_state.get("page_size", 100) in [25, 50, 100, 250, 500]
                else 2,
            )
        )
        table_height = int(
            secondary_controls[1].selectbox(
                "Altura tabela (px)",
                [400, 600, 800, 1000],
                index=[400, 600, 800, 1000].index(table_state.get("table_height", 600))
                if table_state.get("table_height", 600) in [400, 600, 800, 1000]
                else 1,
            )
        )
        auto_width = bool(
            secondary_controls[2].checkbox(
                "Auto largura", value=table_state.get("auto_width", True)
            )
        )
        default_width_profile, width_bucket = _resolve_width_profile_for_bucket(
            table_state
        )
        width_profile = str(
            secondary_controls[3].selectbox(
                "Perfil largura",
                WIDTH_PROFILE_OPTIONS,
                index=WIDTH_PROFILE_OPTIONS.index(default_width_profile),
            )
        )
        _remember_width_profile_for_bucket(table_state, width_bucket, width_profile)

        table_view_df = _build_table_with_derivada_context(view_df, filtered_df)
        if sort_column != "(Sem ordenacao)" and sort_column in table_view_df.columns:
            try:
                table_view_df = table_view_df.sort_values(
                    by=sort_column,
                    ascending=not sort_desc,
                    kind="stable",
                )
            except Exception as exc:
                logger.warning("Falha ao ordenar por %s: %s", sort_column, exc)

        page_size, guarded_page_size = _apply_large_page_guard(
            page_size,
            len(table_view_df),
        )
        if guarded_page_size:
            st.caption(
                "Guard ativo para pagina grande: limite de 500 linhas por pagina."
            )

        _, total_pages = _paginate_dataframe(table_view_df, page=1, page_size=page_size)
        default_page = min(max(int(table_state.get("page_number", 1)), 1), total_pages)
        page_col_left, page_col_right = st.columns([1, 4])
        page_number = int(
            page_col_left.number_input(
                f"Pagina (1..{total_pages})",
                min_value=1,
                max_value=total_pages,
                value=default_page,
                step=1,
            )
        )
        if not compact_mode:
            page_col_right.caption(
                "Dica: use perfil de largura maior para reduzir truncamento de descricao."
            )
        page_df, total_pages = _paginate_dataframe(
            table_view_df, page=page_number, page_size=page_size
        )
        table_state.update(
            {
                "sort_column": sort_column,
                "sort_desc": sort_desc,
                "page_size": page_size,
                "table_height": table_height,
                "auto_width": auto_width,
                "page_number": page_number,
                "width_profile": width_profile,
                "width_profile_by_bucket": table_state.get(
                    "width_profile_by_bucket", {}
                ),
                "table_mode": table_mode,
                "compact_mode": compact_mode,
            }
        )
        _persist_streamlit_state(
            width_profile=str(table_state.get("width_profile", "Padrao (1600)")),
            width_profile_by_bucket=_normalize_width_profile_memory(
                table_state.get("width_profile_by_bucket", {})
            ),
            streamlit_render_stats=_normalize_render_stats(
                st.session_state.get("streamlit_render_stats", {})
            ),
        )

        display_df = ensure_arrow_compatible(page_df.rename(columns=rename_map))
        render_height = _compute_table_render_height(
            page_len=len(page_df),
            configured_height=table_height,
        )
        column_config = _build_streamlit_column_config(
            page_df,
            rename_map,
            available_width=WIDTH_PROFILE_PIXELS.get(width_profile, 1600),
        )

        render_t0 = time.perf_counter()
        st.dataframe(
            display_df,
            width="stretch" if auto_width else "content",
            height=render_height,
            column_config=column_config,
            hide_index=True,
        )
        render_ms = (time.perf_counter() - render_t0) * 1000.0
        _update_render_telemetry(width_profile, render_ms)
        table_caption = _build_table_caption(
            compact_mode=compact_mode,
            page_number=page_number,
            total_pages=total_pages,
            page_len=len(page_df),
            filtered_len=len(table_view_df),
            render_ms=render_ms,
        )
        if not compact_mode:
            st.caption(table_caption)
        else:
            st.caption(table_caption)
        if active_summary and not compact_mode:
            st.markdown("**Filtros ativos:** " + " | ".join(active_summary))
        if "numero_ssa" in filtered_df.columns and "derivada_de" in filtered_df.columns:
            has_derivada, has_children, _ = _compute_derivada_flags(filtered_df)
            chips = (
                f"<span class='ssa-chip'>Derivadas: {int(has_derivada.sum())}</span>"
                f"<span class='ssa-chip'>Com derivadas: {int(has_children.sum())}</span>"
            )
            st.markdown(chips, unsafe_allow_html=True)

        if (
            table_mode == "Tabela + grafico"
            and "situacao" in filtered_df.columns
            and not filtered_df.empty
        ):
            chart_df = (
                filtered_df["situacao"]
                .value_counts()
                .rename_axis("Situacao")
                .reset_index(name="Quantidade")
            )
            st.subheader("Distribuicao por Situacao")
            st.bar_chart(chart_df.set_index("Situacao"))
            extra_charts = st.columns(2)
            if "setor_executor" in filtered_df.columns:
                top_exec = (
                    filtered_df["setor_executor"]
                    .fillna("(vazio)")
                    .astype(str)
                    .value_counts()
                    .head(8)
                    .rename_axis("Setor executor")
                    .reset_index(name="Quantidade")
                )
                extra_charts[0].caption("Principais executores")
                extra_charts[0].bar_chart(top_exec.set_index("Setor executor"))
            if "setor_emissor" in filtered_df.columns:
                top_emis = (
                    filtered_df["setor_emissor"]
                    .fillna("(vazio)")
                    .astype(str)
                    .value_counts()
                    .head(8)
                    .rename_axis("Setor emissor")
                    .reset_index(name="Quantidade")
                )
                extra_charts[1].caption("Principais emissores")
                extra_charts[1].bar_chart(top_emis.set_index("Setor emissor"))

            timeline_charts = st.columns(2)
            if "data_execucao" in filtered_df.columns:
                exec_timeline = pd.to_datetime(
                    filtered_df["data_execucao"], errors="coerce"
                ).dropna()
                if not exec_timeline.empty:
                    exec_weekly = (
                        exec_timeline.dt.to_period("W")
                        .astype(str)
                        .value_counts()
                        .sort_index()
                        .rename_axis("Semana")
                        .reset_index(name="Quantidade")
                    )
                    timeline_charts[0].caption("Execucao por semana")
                    timeline_charts[0].line_chart(exec_weekly.set_index("Semana"))
            if "data_emissao" in filtered_df.columns:
                emit_timeline = pd.to_datetime(
                    filtered_df["data_emissao"], errors="coerce"
                ).dropna()
                if not emit_timeline.empty:
                    emit_weekly = (
                        emit_timeline.dt.to_period("W")
                        .astype(str)
                        .value_counts()
                        .sort_index()
                        .rename_axis("Semana")
                        .reset_index(name="Quantidade")
                    )
                    timeline_charts[1].caption("Emissao por semana")
                    timeline_charts[1].line_chart(emit_weekly.set_index("Semana"))

    with tab_export:
        st.subheader("Exportacao")
        export_left, export_right = st.columns([2.2, 1.8])
        csv_data = view_df.to_csv(index=False).encode("utf-8")
        with export_left:
            st.caption("Arquivos")
            st.download_button(
                "Baixar CSV",
                csv_data,
                file_name=f"ssas_filtradas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv",
                help=f"Exporta {len(view_df)} registros em CSV",
            )
            json_text = view_df.to_json(orient="records", date_format="iso", indent=2)
            if json_text is None:
                raise RuntimeError("to_json retornou None")
            st.download_button(
                "Baixar JSON",
                json_text.encode("utf-8"),
                file_name=f"ssas_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                help="Formato JSON para integracao com APIs",
            )

        with export_right:
            st.caption("Geracao e resumo")
            if st.button("Gerar Excel", help="Prepara arquivo Excel com formatacao"):
                try:
                    import io

                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill

                    buffer = io.BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "SSAs Filtradas"
                    for col_num, col_name in enumerate(view_df.columns, 1):
                        cell = ws.cell(
                            row=1,
                            column=col_num,
                            value=rename_map.get(col_name, col_name),
                        )
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid"
                        )
                    for row_num, row_data in enumerate(
                        view_df.itertuples(index=False), 2
                    ):
                        for col_num, value in enumerate(row_data, 1):
                            ws.cell(row=row_num, column=col_num, value=value)
                    wb.save(buffer)
                    buffer.seek(0)
                    st.download_button(
                        "Excel formatado",
                        buffer.getvalue(),
                        file_name=f"ssas_filtradas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except ImportError:
                    st.error("openpyxl nao instalado - usando CSV simples")
            if st.button("Resumo estatistico", help="Mostra resumo estatistico"):
                stats_info = {
                    "total_registros": len(view_df),
                    "colunas_selecionadas": len(selected_columns),
                    "filtros_ativos": len([x for x in active_summary if x]),
                    "cache_hit_rate": f"{filter_cache.get_stats()['hit_rate']:.1f}%",
                }
                st.json(stats_info)

    with tab_ops:
        ops_left, ops_right = st.columns([1.3, 2.7])
        with ops_left:
            st.subheader("Cache")
            cache_stats = filter_cache.get_stats()
            st.metric("Entradas", f"{cache_stats['size']} / {cache_stats['max_size']}")
            st.metric("Hit rate", f"{cache_stats['hit_rate']:.1f}%")
            st.metric("Evictions", cache_stats["evictions"])
            if st.button("Limpar cache", key="clear_cache_ops"):
                filter_cache.clear()
                st.info("Cache limpo.")
            if hasattr(st, "session_state") and st.session_state is not None:
                render_stats = st.session_state.get("streamlit_render_stats", {})
                if render_stats:
                    profile_options = sorted(str(key) for key in render_stats.keys())
                    default_profile = table_state.get("width_profile", "Padrao (1600)")
                    selected_profile = st.selectbox(
                        "Perfil da telemetria",
                        profile_options,
                        index=profile_options.index(default_profile)
                        if default_profile in profile_options
                        else 0,
                        key="render_telemetry_profile",
                    )
                    if st.button("Limpar telemetria", key="clear_render_telemetry"):
                        st.session_state["streamlit_render_stats"] = {}
                        _persist_streamlit_state(
                            width_profile=str(
                                table_state.get("width_profile", "Padrao (1600)")
                            ),
                            width_profile_by_bucket=_normalize_width_profile_memory(
                                table_state.get("width_profile_by_bucket", {})
                            ),
                            streamlit_render_stats={},
                        )
                        st.info("Telemetria limpa.")
                    profile_stats = render_stats.get(selected_profile)
                else:
                    profile_stats = None
                if profile_stats:
                    st.caption(
                        _format_render_stats_line(selected_profile, profile_stats)
                    )

        with ops_right:
            with st.expander("Fonte de dados avancada", expanded=False):
                source_db_input = st.text_input(
                    "Arquivo do banco",
                    value=db_path,
                    key="ops_source_db_path",
                )
                source_docs_input = st.text_input(
                    "Pasta com planilhas",
                    value=docs_dir,
                    key="ops_source_docs_dir",
                )
                source_actions = st.columns([1.1, 1.1, 1.8])
                apply_source = source_actions[0].button(
                    "Aplicar fonte", key="apply_source_paths"
                )
                run_load = source_actions[1].button(
                    "Carregar dados", key="load_data_ops"
                )
                run_reimport = source_actions[2].button(
                    "Reimportar planilhas", key="reimport_data_ops"
                )
                if apply_source:
                    try:
                        resolved_db = str(
                            ensure_path_is_allowed(
                                source_db_input,
                                purpose="Arquivo do banco",
                                expect_directory=False,
                            )
                        )
                        resolved_docs = str(
                            ensure_path_is_allowed(
                                source_docs_input,
                                purpose="Pasta com planilhas",
                                expect_directory=True,
                            )
                        )
                        st.session_state["streamlit_source_state"] = {
                            "db_path": resolved_db,
                            "docs_dir": resolved_docs,
                        }
                        st.success(
                            "Fonte aplicada. Recarregue dados para refletir mudancas."
                        )
                    except PathSafetyError as exc:
                        st.error(str(exc))
                if run_load or run_reimport:
                    source_state = st.session_state.get("streamlit_source_state", {})
                    op_db = str(source_state.get("db_path", db_path))
                    op_docs = str(source_state.get("docs_dir", docs_dir))
                    try:
                        ok = import_files_to_database(
                            docs_dir=op_docs,
                            db_path=op_db,
                            force_import=bool(run_reimport),
                            raise_on_error=True,
                        )
                        if hasattr(load_dataframe, "clear"):
                            load_dataframe.clear()
                        filter_cache.clear()
                        _clear_recent_api_snapshot()
                        if ok:
                            st.success("Importacao concluida.")
                        else:
                            st.info("Nenhum arquivo novo processado.")
                        rerun_fn = getattr(st, "rerun", None)
                        if callable(rerun_fn):
                            rerun_fn()
                    except Exception as exc:
                        st.error(f"Importacao falhou: {exc}")
            st.subheader("API Itaipu")
            if consult_api:
                if hasattr(st, "session_state") and st.session_state is not None:
                    if "recent_api_df" not in st.session_state:
                        st.session_state["recent_api_df"] = None
                api_actions = st.columns([1.2, 1.2, 2.6])
                if api_actions[0].button("Atualizar dados API", key="refresh_api_data"):
                    try:
                        api_items = fetch_pending_ssas(
                            years=1,
                            opts=RequestOptions(timeout=_resolve_api_timeout_seconds()),
                        )
                        mapped = map_to_dataframe(api_items)
                        if mapped is not None and not mapped.empty:
                            cols = [
                                col
                                for col in (
                                    "numero_ssa",
                                    "situacao",
                                    "setor_emissor",
                                    "setor_executor",
                                    "descricao_ssa",
                                    "data_cadastro",
                                )
                                if col in mapped.columns
                            ]
                            recent_df = mapped[cols].copy()
                            if (
                                hasattr(st, "session_state")
                                and st.session_state is not None
                            ):
                                st.session_state["recent_api_df"] = recent_df
                            st.success(f"API retornou {len(recent_df)} registros.")
                        else:
                            st.info("API sem novos registros.")
                    except Exception as exc:  # noqa: BLE001
                        logger.error("Falha ao consultar API Itaipu: %s", exc)
                        st.warning(
                            "Nao foi possivel consultar API. Dashboard segue com base local."
                        )
                if api_actions[1].button("Limpar snapshot API", key="clear_api_data"):
                    _clear_recent_api_snapshot()
                    st.info("Snapshot de API removido.")
                api_actions[2].caption(
                    "Atualizacao manual para evitar bloqueio em reruns."
                )
                if hasattr(st, "session_state") and st.session_state is not None:
                    recent_df = st.session_state.get("recent_api_df")
                if _api_snapshot_available(consult_api, recent_df):
                    snapshot_df = cast(pd.DataFrame, recent_df)
                    st.dataframe(
                        ensure_arrow_compatible(snapshot_df),
                        width="stretch",
                        height=240,
                    )
            else:
                st.info(
                    "Ative a opcao de API na aba Filtros para consultar dados recentes."
                )
