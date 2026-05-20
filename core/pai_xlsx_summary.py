"""Streaming summary reader for normalized PAI XLSX files."""

from __future__ import annotations

from dataclasses import dataclass, field
import logging
from pathlib import Path
from typing import Any, TypedDict

import pandas as pd
from openpyxl import load_workbook

SUMMARY_SSA_EXAMPLE_LIMIT = 20
SUMMARY_GROUP_EXAMPLE_LIMIT = 10
logger = logging.getLogger(__name__)


class PaiXlsxSummary(TypedDict):
    ssa_examples: list[str]
    rows_by_executor_sector: dict[str, int]
    rows_by_emitter_sector: dict[str, int]
    rows_by_source_file: dict[str, int]
    ssa_examples_by_executor_sector: dict[str, list[str]]
    summary_error: str | None


@dataclass
class _PaiXlsxSummaryAccumulator:
    ssa_examples: list[str] = field(default_factory=list)
    rows_by_executor_sector: dict[str, int] = field(default_factory=dict)
    rows_by_emitter_sector: dict[str, int] = field(default_factory=dict)
    rows_by_source_file: dict[str, int] = field(default_factory=dict)
    ssa_examples_by_executor_sector: dict[str, list[str]] = field(default_factory=dict)

    def add_row(self, row_values: dict[str, str]) -> None:
        number = row_values.get("number", "")
        executor = row_values.get("executor", "")
        emitter = row_values.get("emitter", "")
        source_file = row_values.get("source_file", "")
        if number and len(self.ssa_examples) < SUMMARY_SSA_EXAMPLE_LIMIT:
            self.ssa_examples.append(number)
        if executor:
            self.rows_by_executor_sector[executor] = (
                self.rows_by_executor_sector.get(executor, 0) + 1
            )
        if emitter:
            self.rows_by_emitter_sector[emitter] = (
                self.rows_by_emitter_sector.get(emitter, 0) + 1
            )
        if source_file:
            self.rows_by_source_file[source_file] = (
                self.rows_by_source_file.get(source_file, 0) + 1
            )
        if executor and number:
            bucket = self.ssa_examples_by_executor_sector.setdefault(executor, [])
            if len(bucket) < SUMMARY_GROUP_EXAMPLE_LIMIT:
                bucket.append(number)

    def to_summary(self) -> PaiXlsxSummary:
        return PaiXlsxSummary(
            ssa_examples=self.ssa_examples,
            rows_by_executor_sector=self.rows_by_executor_sector,
            rows_by_emitter_sector=self.rows_by_emitter_sector,
            rows_by_source_file=self.rows_by_source_file,
            ssa_examples_by_executor_sector=self.ssa_examples_by_executor_sector,
            summary_error=None,
        )


def empty_pai_xlsx_summary() -> PaiXlsxSummary:
    return _empty_pai_xlsx_summary(None)


def summarize_normalized_pai_frame(frame: pd.DataFrame) -> PaiXlsxSummary:
    number_values = _normalized_text_series(frame, "numero_ssa")
    executor_values = _normalized_text_series(frame, "setor_executor")
    emitter_values = _normalized_text_series(frame, "setor_emissor")
    source_file_values = _normalized_text_series(frame, "arquivo_origem")
    grouped_examples = pd.DataFrame(
        {"executor": executor_values, "number": number_values}
    )
    examples_by_executor: dict[str, list[str]] = {
        executor: group["number"].head(SUMMARY_GROUP_EXAMPLE_LIMIT).tolist()
        for executor, group in grouped_examples[
            (grouped_examples["executor"] != "") & (grouped_examples["number"] != "")
        ].groupby("executor", sort=False)
    }
    return PaiXlsxSummary(
        ssa_examples=number_values[number_values != ""]
        .head(SUMMARY_SSA_EXAMPLE_LIMIT)
        .tolist(),
        rows_by_executor_sector=executor_values[executor_values != ""]
        .value_counts(sort=False)
        .to_dict(),
        rows_by_emitter_sector=emitter_values[emitter_values != ""]
        .value_counts(sort=False)
        .to_dict(),
        rows_by_source_file=source_file_values[source_file_values != ""]
        .value_counts(sort=False)
        .to_dict(),
        ssa_examples_by_executor_sector=examples_by_executor,
        summary_error=None,
    )


def _empty_pai_xlsx_summary(summary_error: str | None) -> PaiXlsxSummary:
    return PaiXlsxSummary(
        ssa_examples=[],
        rows_by_executor_sector={},
        rows_by_emitter_sector={},
        rows_by_source_file={},
        ssa_examples_by_executor_sector={},
        summary_error=summary_error,
    )


def read_pai_xlsx_summary(path: Path | None) -> PaiXlsxSummary:
    if path is None or not path.is_file():
        return empty_pai_xlsx_summary()
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError) as exc:
        logger.warning("Failed to read PAI summary XLSX %s: %s", path, exc)
        return _summary_read_error(type(exc).__name__)
    try:
        sheet = workbook.active
        if sheet is None:
            return _empty_pai_xlsx_summary("summary_read_error:no_active_sheet")
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            return _empty_pai_xlsx_summary("summary_read_error:missing_header")
        column_positions = _summary_column_positions(header_row)
        if all(position is None for position in column_positions.values()):
            return _empty_pai_xlsx_summary("summary_read_error:no_summary_columns")
        active_positions = {
            key: position
            for key, position in column_positions.items()
            if position is not None
        }
        summary = _PaiXlsxSummaryAccumulator()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            summary.add_row(
                {
                    key: _row_value(row, position)
                    for key, position in active_positions.items()
                }
            )
        return summary.to_summary()
    finally:
        if workbook is not None:
            workbook.close()


def _summary_read_error(error_name: str) -> PaiXlsxSummary:
    return _empty_pai_xlsx_summary(f"summary_read_error:{error_name}")


def _normalized_text_series(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series("", index=frame.index)
    return frame[column].fillna("").astype("string").str.strip()


def _summary_column_positions(header_row: tuple[Any, ...]) -> dict[str, int | None]:
    header_values = ["" if value is None else str(value).strip() for value in header_row]
    header_lookup = {value.casefold(): index for index, value in enumerate(header_values)}
    return {
        key: _column_position(header_lookup, aliases)
        for key, aliases in {
            "number": ("numero_ssa", "Numero da SSA", "ssa_number", "SSA"),
            "executor": ("setor_executor", "executor_sector", "Setor Executor"),
            "emitter": ("setor_emissor", "emitter_sector", "Setor Emissor"),
            "source_file": ("arquivo_origem", "source_file", "Arquivo Origem"),
        }.items()
    }


def _column_position(
    header_lookup: dict[str, int],
    candidates: tuple[str, ...],
) -> int | None:
    for candidate in candidates:
        found = header_lookup.get(candidate.casefold())
        if found is not None:
            return found
    return None


def _row_value(row: tuple[Any, ...], position: int | None) -> str:
    if position is None or position >= len(row) or row[position] is None:
        return ""
    return str(row[position]).strip()
